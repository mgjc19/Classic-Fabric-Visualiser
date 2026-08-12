"""
Classic Fabric Visualiser - Backend API
Handles file uploads, parsing, and topology construction.
"""
import io
import json
import os
import re
import zipfile
import tempfile
import httpx
from pathlib import Path

from fastapi import FastAPI, UploadFile, File, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from starlette.responses import Response, StreamingResponse, JSONResponse

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

from parsers import (
    parse_cdp_neighbors,
    parse_lldp_neighbors,
    parse_interface_brief,
    parse_interface_description,
    parse_interface_status,
    parse_platform_detail,
    parse_running_config,
    extract_hostname,
    parse_bgp_summary,
    parse_bgp_neighbors_detail,
    parse_bgp_from_config,
    parse_ospf_overview,
    parse_ospf_neighbors,
    parse_ospf_interfaces,
    parse_ospf_from_config,
)
from parsers.interface_parser import infer_neighbors_from_descriptions
from topology import TopologyBuilder
from topology.routing_builder import RoutingTopologyBuilder
from migration import MigrationClassifier, UnderlayDesigner
from fabric_builder import BomParser, FabricModel, ConfigEngine, YamlExporter, NxosExporter, EndpointStore, TrafficEngine, FailoverSimulator

app = FastAPI(
    title="Classic Fabric Visualiser",
    description="Visualise traditional ethernet fabric topologies",
    version="1.0.0",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

FRONTEND_DIR = Path(__file__).parent.parent / "frontend"
UPLOAD_MAX_SIZE = 100 * 1024 * 1024  # 100MB per file


@app.post("/api/upload")
async def upload_files(files: list[UploadFile] = File(...)):
    """
    Upload device output files and build topology.
    Accepts .txt, .log, .zip, .xlsx files.
    """
    all_texts: list[tuple[str, str]] = []

    for upload_file in files:
        if not upload_file.filename:
            continue

        content = await upload_file.read()
        if len(content) > UPLOAD_MAX_SIZE:
            raise HTTPException(
                status_code=413,
                detail=f"File {upload_file.filename} exceeds 100MB limit"
            )

        filename_lower = upload_file.filename.lower()

        if filename_lower.endswith(".zip"):
            extracted = _extract_zip(content)
            all_texts.extend(extracted)
        elif filename_lower.endswith((".xlsx", ".xls")):
            extracted = _extract_excel(content)
            all_texts.extend(extracted)
        else:
            text = content.decode("utf-8", errors="replace")
            all_texts.append((upload_file.filename, text))

    if not all_texts:
        raise HTTPException(status_code=400, detail="No valid files found")

    topology = _process_files(all_texts)
    return topology


@app.post("/api/upload-stream")
async def upload_files_stream(files: list[UploadFile] = File(...)):
    """
    Upload files and stream parsing progress as Server-Sent Events.
    The final event contains the complete topology JSON.
    """
    all_texts: list[tuple[str, str]] = []

    for upload_file in files:
        if not upload_file.filename:
            continue

        content = await upload_file.read()
        if len(content) > UPLOAD_MAX_SIZE:
            raise HTTPException(
                status_code=413,
                detail=f"File {upload_file.filename} exceeds 100MB limit"
            )

        filename_lower = upload_file.filename.lower()

        if filename_lower.endswith(".zip"):
            extracted = _extract_zip(content)
            all_texts.extend(extracted)
        elif filename_lower.endswith((".xlsx", ".xls")):
            extracted = _extract_excel(content)
            all_texts.extend(extracted)
        else:
            text = content.decode("utf-8", errors="replace")
            all_texts.append((upload_file.filename, text))

    if not all_texts:
        raise HTTPException(status_code=400, detail="No valid files found")

    async def event_generator():
        import asyncio

        def emit(action: str, message: str):
            payload = json.dumps({"action": action, "message": message})
            return f"data: {payload}\n\n"

        yield emit("upload", f"Processing {len(all_texts)} file(s) from upload...")

        builder = TopologyBuilder()
        device_files: dict[str, list[str]] = {}

        for filename, text in all_texts:
            yield emit("file", f"Reading {filename}")
            await asyncio.sleep(0)

            hostname = _detect_hostname(filename, text)
            if hostname:
                if hostname not in device_files:
                    device_files[hostname] = []
                    yield emit("device", f"Discovered device: {hostname}")
                device_files[hostname].append(text)
            else:
                sections = _split_multi_device_output(text)
                if sections:
                    for host, section_text in sections:
                        if host not in device_files:
                            device_files[host] = []
                            yield emit("device", f"Discovered device: {host}")
                        device_files[host].append(section_text)
                else:
                    fallback_name = Path(filename).stem
                    if fallback_name not in device_files:
                        device_files[fallback_name] = []
                        yield emit("device", f"Discovered device: {fallback_name}")
                    device_files[fallback_name].append(text)

        yield emit("parse", f"Found {len(device_files)} devices. Parsing command outputs...")
        await asyncio.sleep(0)

        routing_builder = RoutingTopologyBuilder()
        device_count = 0
        total_devices = len(device_files)
        known_hostnames = set(device_files.keys())

        for hostname, texts in device_files.items():
            device_count += 1
            yield emit("parse", f"[{device_count}/{total_devices}] Parsing {hostname}...")
            await asyncio.sleep(0)

            combined_text = "\n".join(texts)
            commands_found = _parse_device_output(builder, hostname, combined_text, routing_builder, known_hostnames)
            if commands_found:
                yield emit("commands", f"  Found: {', '.join(commands_found)}")
                await asyncio.sleep(0)

        yield emit("parse", "Building topology graphs...")
        await asyncio.sleep(0)

        topology = builder.build()
        topology["devices_parsed"] = list(device_files.keys())

        routing_builder.set_physical_devices(builder.devices)
        bgp_topology = routing_builder.build_bgp_topology()
        ospf_topology = routing_builder.build_ospf_topology()

        topology["bgp"] = bgp_topology
        topology["ospf"] = ospf_topology

        topology["migration"] = _build_migration_data(topology)

        stats = topology["stats"]
        bgp_count = bgp_topology["stats"]["total_peers"]
        ospf_count = ospf_topology["stats"]["total_adjacencies"]
        summary = f"Topology ready: {stats['total_devices']} devices, {stats['total_links']} links"
        if bgp_count > 0:
            summary += f", {bgp_count} BGP peers"
        if ospf_count > 0:
            summary += f", {ospf_count} OSPF adjacencies"
        yield emit("done", summary)

        yield f"event: topology\ndata: {json.dumps(topology)}\n\n"

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        }
    )


@app.get("/api/health")
async def health():
    return {"status": "ok", "version": "1.0.0"}


if FRONTEND_DIR.exists():
    from starlette.responses import Response

    @app.get("/")
    async def serve_index():
        content = (FRONTEND_DIR / "index.html").read_bytes()
        return Response(content, media_type="text/html", headers={"Cache-Control": "no-store"})

    @app.get("/test")
    async def serve_test():
        content = (FRONTEND_DIR / "test.html").read_bytes()
        return Response(content, media_type="text/html", headers={"Cache-Control": "no-store"})

    @app.get("/static/{filepath:path}")
    async def serve_static(filepath: str):
        file_path = FRONTEND_DIR / filepath
        if not file_path.exists() or not file_path.is_file():
            raise HTTPException(status_code=404)
        content = file_path.read_bytes()
        media_types = {
            ".js": "application/javascript",
            ".css": "text/css",
            ".html": "text/html",
            ".png": "image/png",
            ".svg": "image/svg+xml",
        }
        ext = file_path.suffix.lower()
        mt = media_types.get(ext, "application/octet-stream")
        return Response(content, media_type=mt, headers={"Cache-Control": "no-store"})


def _extract_zip(content: bytes) -> list[tuple[str, str]]:
    """Extract text files from a ZIP archive."""
    texts = []
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as zf:
            for name in zf.namelist():
                if name.startswith("__MACOSX") or name.startswith("."):
                    continue
                name_lower = name.lower()
                if name_lower.endswith((".txt", ".log", ".cfg", ".conf", "")):
                    try:
                        data = zf.read(name)
                        if len(data) > UPLOAD_MAX_SIZE:
                            continue
                        text = data.decode("utf-8", errors="replace")
                        if text.strip():
                            texts.append((name, text))
                    except (KeyError, RuntimeError):
                        continue
    except zipfile.BadZipFile:
        pass
    return texts


def _extract_excel(content: bytes) -> list[tuple[str, str]]:
    """Extract text content from Excel files (each sheet as a separate device)."""
    texts = []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            lines = []
            for row in ws.iter_rows(values_only=True):
                row_text = "\t".join(str(cell) if cell is not None else "" for cell in row)
                if row_text.strip():
                    lines.append(row_text)
            if lines:
                texts.append((sheet_name, "\n".join(lines)))
        wb.close()
    except Exception:
        pass
    return texts


def _process_files(file_texts: list[tuple[str, str]]) -> dict:
    """
    Process all uploaded file contents and build the topology.
    Auto-detects command outputs and parses accordingly.
    """
    builder = TopologyBuilder()
    device_files: dict[str, list[str]] = {}
    parse_log: list[dict] = []

    for filename, text in file_texts:
        parse_log.append({"action": "file", "message": f"Reading {filename}"})
        hostname = _detect_hostname(filename, text)
        if hostname:
            if hostname not in device_files:
                device_files[hostname] = []
                parse_log.append({"action": "device", "message": f"Discovered device: {hostname}"})
            device_files[hostname].append(text)
        else:
            sections = _split_multi_device_output(text)
            if sections:
                for host, section_text in sections:
                    if host not in device_files:
                        device_files[host] = []
                        parse_log.append({"action": "device", "message": f"Discovered device: {host}"})
                    device_files[host].append(section_text)
            else:
                fallback_name = Path(filename).stem
                if fallback_name not in device_files:
                    device_files[fallback_name] = []
                    parse_log.append({"action": "device", "message": f"Discovered device: {fallback_name}"})
                device_files[fallback_name].append(text)

    routing_builder2 = RoutingTopologyBuilder()
    known_hostnames = set(device_files.keys())
    for hostname, texts in device_files.items():
        parse_log.append({"action": "parse", "message": f"Parsing {hostname}..."})
        combined_text = "\n".join(texts)
        commands_found = _parse_device_output(builder, hostname, combined_text, routing_builder2, known_hostnames)
        if commands_found:
            parse_log.append({"action": "commands", "message": f"  Found: {', '.join(commands_found)}"})

    topology = builder.build()
    topology["devices_parsed"] = list(device_files.keys())
    topology["parse_log"] = parse_log

    routing_builder2.set_physical_devices(builder.devices)
    topology["bgp"] = routing_builder2.build_bgp_topology()
    topology["ospf"] = routing_builder2.build_ospf_topology()

    topology["migration"] = _build_migration_data(topology)

    parse_log.append({"action": "done", "message": f"Built topology: {topology['stats']['total_devices']} devices, {topology['stats']['total_links']} links"})

    return topology


def _build_migration_data(topology: dict) -> dict:
    """Run migration classification and initial underlay design."""
    try:
        classifier = MigrationClassifier(topology)
        classifications = classifier.classify_all()
        phases = classifier.suggest_phases(classifications)
        vni_mapping = classifier.generate_vni_mapping()

        nodes_map = {n["data"]["id"]: n["data"] for n in topology.get("nodes", [])}
        adjacency: dict[str, list[dict]] = {}
        for edge in topology.get("edges", []):
            src = edge["data"]["source"]
            tgt = edge["data"]["target"]
            adjacency.setdefault(src, []).append(edge["data"])
            adjacency.setdefault(tgt, []).append(edge["data"])

        designer = UnderlayDesigner(classifications, nodes_map, adjacency)
        underlay_design = designer.design(
            underlay_protocol="ospf",
            bgp_afs=["l2vpn_evpn"],
        )

        return {
            "classifications": classifications,
            "phases": phases,
            "vni_mapping": vni_mapping,
            "underlay_design": underlay_design,
        }
    except Exception:
        return {"classifications": {}, "phases": [], "vni_mapping": [], "underlay_design": {}}


_last_topology: dict = {}


@app.post("/api/redesign-underlay")
async def redesign_underlay(request: Request):
    """Re-compute underlay/overlay design with user-selected parameters."""
    body = await request.json()

    underlay_protocol = body.get("underlay_protocol", "ospf")
    bgp_afs = body.get("bgp_afs", ["l2vpn_evpn"])
    ospf_area = body.get("ospf_area", "0.0.0.0")
    spine_asn = int(body.get("spine_asn", 65000))
    leaf_asn_start = int(body.get("leaf_asn_start", 65001))
    overlay_asn = body.get("overlay_asn")
    if overlay_asn is not None:
        overlay_asn = int(overlay_asn)

    classifications = body.get("classifications", {})
    nodes = body.get("nodes", {})
    adjacency = body.get("adjacency", {})

    if not classifications:
        return JSONResponse({"error": "No classification data provided"}, status_code=400)

    designer = UnderlayDesigner(classifications, nodes, adjacency)
    result = designer.design(
        underlay_protocol=underlay_protocol,
        bgp_afs=bgp_afs,
        ospf_area=ospf_area,
        spine_asn=spine_asn,
        leaf_asn_start=leaf_asn_start,
        overlay_asn=overlay_asn,
    )
    return result


# =============================================================================
# Fabric Builder API (Phase 4)
# =============================================================================

_fabric_model: FabricModel | None = None
_fabric_configs: dict[str, str] = {}
_endpoint_store: EndpointStore = EndpointStore()
_traffic_engine: TrafficEngine | None = None
_failover_sim: FailoverSimulator | None = None


def _init_traffic_engine():
    """Initialize or reinitialize traffic engine when fabric model changes."""
    global _traffic_engine, _failover_sim
    if _fabric_model:
        _traffic_engine = TrafficEngine(_fabric_model, _endpoint_store)
        _failover_sim = FailoverSimulator(_fabric_model, _endpoint_store, _traffic_engine)


@app.post("/api/fabric/upload-bom")
async def upload_bom(file: UploadFile = File(...)):
    """Upload BOM (Excel/CSV) and parse into a fabric model.
    Supports hardware BOMs (PIDs/quantities) and fabric BOMs (hostnames/IPs).
    For hardware BOMs, returns the parsed inventory; call /api/fabric/build-from-hardware to generate devices.
    """
    global _fabric_model, _fabric_configs

    if not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")

    content = await file.read()
    if len(content) > UPLOAD_MAX_SIZE:
        raise HTTPException(status_code=413, detail="File exceeds 100MB limit")

    parser = BomParser()
    try:
        bom_data = parser.parse(content, file.filename)
    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f"Failed to parse BOM file: {str(e)}"
        )

    bom_type = bom_data.get("type", "hardware")

    if bom_type == "hardware":
        hardware = bom_data.get("hardware", {})
        if not hardware or not hardware.get("switches"):
            raise HTTPException(
                status_code=400,
                detail="No Nexus switches (N9K-*) found in BOM. Ensure your file contains Cisco PIDs."
            )
        return {
            "type": "hardware",
            "hardware": hardware,
            "metadata": bom_data["metadata"],
        }

    if not bom_data["devices"]:
        raise HTTPException(
            status_code=400,
            detail="No devices found in BOM file. Ensure your file has a column with device hostnames."
        )

    _fabric_model = FabricModel()
    _fabric_model.load_from_bom(bom_data)
    _fabric_model.add_default_overlay()
    _fabric_configs = {}
    _init_traffic_engine()

    return {"type": "fabric", **_fabric_model.to_dict()}


@app.post("/api/fabric/build-from-hardware")
async def build_from_hardware(request: Request):
    """Generate a fabric model from a parsed hardware BOM with user-specified naming/IP config.
    Supports multi-site: pass 'sites' array to split inventory across multiple sites.
    """
    global _fabric_model, _fabric_configs

    body = await request.json()
    hardware = body.get("hardware")
    sites = body.get("sites", [])
    config = body.get("config")

    print(f"[build-from-hardware] received {len(sites)} sites, config={'yes' if config else 'no'}")
    print(f"[build-from-hardware] site names = {[s.get('site','?') for s in sites]}")

    if not hardware or not hardware.get("switches"):
        raise HTTPException(status_code=400, detail="No hardware data provided")

    if not sites and config:
        sites = [config]

    if not sites:
        sites = [{"site": "DC1"}]

    print(f"[build-from-hardware] processing {len(sites)} sites after defaults")

    if len(sites) == 1:
        bom_data = BomParser.generate_devices_from_hardware(hardware, sites[0])
    else:
        all_devices = []
        all_links = []
        num_sites = len(sites)

        split_hardware_per_site = []
        for site_idx in range(num_sites):
            site_hw = {"switches": [], "sfps": hardware.get("sfps", []), "cables": hardware.get("cables", [])}
            for sw in hardware.get("switches", []):
                per_site_qty = sw["quantity"] // num_sites
                remainder = sw["quantity"] % num_sites
                qty = per_site_qty + (1 if site_idx < remainder else 0)
                if qty > 0:
                    site_sw = dict(sw)
                    site_sw["quantity"] = qty
                    site_hw["switches"].append(site_sw)
            split_hardware_per_site.append(site_hw)

        for site_idx, site_config in enumerate(sites):
            site_data = BomParser.generate_devices_from_hardware(
                split_hardware_per_site[site_idx], site_config
            )
            all_devices.extend(site_data["devices"])
            all_links.extend(site_data["links"])

        # Generate DCI/inter-site links between border gateways or border leaves
        # DCI peering uses BGP address-family IPv4 unicast between border gateway devices
        import uuid as _uuid
        dci_candidates_by_site: dict[str, list[dict]] = {}
        for dev in all_devices:
            site_name = dev.get("site", "")
            role = dev.get("role", "")
            if role in ("border_gateway", "border_leaf"):
                dci_candidates_by_site.setdefault(site_name, []).append(dev)

        # If no BGW/BLeaf exists, promote the last leaf device per site
        # to border_gateway role (IEEE standard: 1 BGW per site for multi-site VXLAN)
        if not any(dci_candidates_by_site.values()):
            leaves_by_site: dict[str, list[dict]] = {}
            for dev in all_devices:
                if dev.get("role") == "leaf":
                    leaves_by_site.setdefault(dev.get("site", ""), []).append(dev)
            for site_name, site_leaves in leaves_by_site.items():
                promoted = [site_leaves[-1]]
                for dev in promoted:
                    dev["role"] = "border_gateway"
                    old_hostname = dev["hostname"]
                    site_prefix = site_name or "DC1"
                    dev["hostname"] = f"{site_prefix}-BGW-01"
                    for lnk in all_links:
                        if lnk.get("from_device") == old_hostname:
                            lnk["from_device"] = dev["hostname"]
                        if lnk.get("to_device") == old_hostname:
                            lnk["to_device"] = dev["hostname"]
                dci_candidates_by_site[site_name] = promoted

        # Assign a shared DCI BGP ASN for border gateways
        dci_bgp_asn = 65500
        site_names = sorted(dci_candidates_by_site.keys())
        for idx, sn in enumerate(site_names):
            site_asn = dci_bgp_asn + idx
            for dev in dci_candidates_by_site[sn]:
                dev["asn"] = str(site_asn)

        for i in range(len(site_names)):
            for j in range(i + 1, len(site_names)):
                site_a_devs = dci_candidates_by_site[site_names[i]]
                site_b_devs = dci_candidates_by_site[site_names[j]]
                for a_dev in site_a_devs:
                    for b_dev in site_b_devs:
                        dci_link = {
                            "id": str(_uuid.uuid4()),
                            "from_device": a_dev["hostname"],
                            "from_port": "Ethernet1/48",
                            "to_device": b_dev["hostname"],
                            "to_port": "Ethernet1/48",
                            "sfp": "",
                            "cable_type": "DCI",
                            "speed": "100G",
                            "protocol": "BGP",
                            "bgp_address_family": "ipv4 unicast",
                            "from_asn": a_dev.get("asn", ""),
                            "to_asn": b_dev.get("asn", ""),
                        }
                        all_links.append(dci_link)

        all_sites = sorted(set(d.get("site", "") for d in all_devices if d.get("site")))
        bom_data = {
            "type": "fabric",
            "devices": all_devices,
            "links": all_links,
            "hardware": None,
            "metadata": {
                "total_devices": len(all_devices),
                "total_links": len(all_links),
                "sites": all_sites if all_sites else ["site-1"],
                "multisite": len(all_sites) > 1,
                "bom_type": "generated_from_hardware",
            }
        }

    _fabric_model = FabricModel()
    _fabric_model.load_from_bom(bom_data)
    _fabric_model.add_default_overlay()
    _fabric_configs = {}
    _init_traffic_engine()

    return _fabric_model.to_dict()


@app.get("/api/fabric/template")
async def download_bom_template():
    """Download a BOM Excel template."""
    template_bytes = BomParser.generate_template()
    return Response(
        content=template_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=fabric_bom_template.xlsx"}
    )


@app.get("/api/fabric/model")
async def get_fabric_model():
    """Get the current fabric model state."""
    if not _fabric_model:
        raise HTTPException(status_code=404, detail="No fabric model loaded. Upload a BOM first.")
    return _fabric_model.to_dict()


@app.post("/api/fabric/load-demo")
async def load_demo_fabric(request: Request):
    """Load a pre-built demo multi-site VXLAN fabric for immediate use."""
    global _fabric_model, _fabric_configs
    body = await request.json()
    _fabric_model = FabricModel()
    _fabric_model.devices = []
    _fabric_model.links = []
    _fabric_model.sites = body.get("sites", [])
    _fabric_model.multisite = body.get("multisite", False)

    overlay_data = body.get("overlay", {})
    if overlay_data:
        _fabric_model.overlay.vrfs = overlay_data.get("vrfs", [])
        _fabric_model.overlay.vlans = overlay_data.get("vlans", [])
        _fabric_model.overlay.vnis = overlay_data.get("vnis", [])

    gc = body.get("global_config")
    if gc:
        _fabric_model.global_config.update(gc)
    d2 = body.get("day2_config")
    if d2:
        _fabric_model.day2_config.update(d2)

    from fabric_builder.fabric_model import FabricDevice, FabricLink
    for dev in body.get("devices", []):
        _fabric_model.devices.append(FabricDevice(dev))
    for link in body.get("links", []):
        _fabric_model.links.append(FabricLink(link))

    _fabric_configs.clear()
    _endpoint_store.endpoints = []

    for ep in body.get("endpoints", []):
        from fabric_builder.endpoint_model import FabricEndpoint
        _endpoint_store.endpoints.append(FabricEndpoint(ep))

    _init_traffic_engine()
    return {"status": "ok", "devices": len(_fabric_model.devices), "links": len(_fabric_model.links)}


@app.put("/api/fabric/device/{device_id}")
async def update_device(device_id: str, request: Request):
    """Edit device properties."""
    if not _fabric_model:
        raise HTTPException(status_code=404, detail="No fabric model loaded")
    body = await request.json()
    result = _fabric_model.update_device(device_id, body)
    if not result:
        raise HTTPException(status_code=404, detail="Device not found")
    _fabric_configs.clear()
    return result


@app.put("/api/fabric/link/{link_id}")
async def update_link(link_id: str, request: Request):
    """Edit link properties."""
    if not _fabric_model:
        raise HTTPException(status_code=404, detail="No fabric model loaded")
    body = await request.json()
    result = _fabric_model.update_link(link_id, body)
    if not result:
        raise HTTPException(status_code=404, detail="Link not found")
    _fabric_configs.clear()
    return result


@app.post("/api/fabric/overlay")
async def update_overlay(request: Request):
    """Add/edit VRFs, VLANs, VNIs."""
    if not _fabric_model:
        raise HTTPException(status_code=404, detail="No fabric model loaded")
    body = await request.json()
    _fabric_model.update_overlay(body)
    _fabric_configs.clear()
    return _fabric_model.overlay.to_dict()


@app.put("/api/fabric/global-config")
async def update_global_config(request: Request):
    """Update global fabric configuration."""
    if not _fabric_model:
        raise HTTPException(status_code=404, detail="No fabric model loaded")
    body = await request.json()
    for key, value in body.items():
        if key in _fabric_model.global_config:
            _fabric_model.global_config[key] = value
    _fabric_configs.clear()
    return _fabric_model.global_config


@app.put("/api/fabric/day2-config")
async def update_day2_config(request: Request):
    """Update Day-2 configuration (NTP, SNMP, etc.)."""
    if not _fabric_model:
        raise HTTPException(status_code=404, detail="No fabric model loaded")
    body = await request.json()
    for key, value in body.items():
        if key in _fabric_model.day2_config:
            _fabric_model.day2_config[key] = value
    _fabric_configs.clear()
    return _fabric_model.day2_config


@app.post("/api/fabric/generate-config")
async def generate_configs():
    """Generate all configs from the current model."""
    global _fabric_configs
    if not _fabric_model:
        raise HTTPException(status_code=404, detail="No fabric model loaded")
    engine = ConfigEngine(_fabric_model)
    _fabric_configs = engine.generate_all()
    return {"devices": list(_fabric_configs.keys()), "total": len(_fabric_configs)}


@app.get("/api/fabric/config/{device_id}")
async def get_device_config(device_id: str):
    """Get generated config for one device."""
    global _fabric_configs
    if not _fabric_model:
        raise HTTPException(status_code=404, detail="No fabric model loaded")
    if not _fabric_configs:
        engine = ConfigEngine(_fabric_model)
        _fabric_configs = engine.generate_all()
    device = _fabric_model.get_device(device_id)
    if not device:
        raise HTTPException(status_code=404, detail="Device not found")
    config = _fabric_configs.get(device.hostname, "")
    if not config:
        engine = ConfigEngine(_fabric_model)
        config = engine.get_device_config(device_id)
    return {"hostname": device.hostname, "config": config}


@app.post("/api/fabric/cli-command")
async def apply_cli_command(request: Request):
    """Apply a CLI-style command to the model (config terminal concept)."""
    if not _fabric_model:
        raise HTTPException(status_code=404, detail="No fabric model loaded")
    body = await request.json()
    device_id = body.get("device_id", "")
    command = body.get("command", "").strip()

    if not device_id or not command:
        raise HTTPException(status_code=400, detail="device_id and command are required")

    device = _fabric_model.get_device(device_id)
    if not device:
        raise HTTPException(status_code=404, detail="Device not found")

    old_hostname = device.hostname
    ctx = _cli_contexts.get(device_id, {"mode": "config", "sub": None, "target": None})
    result, new_ctx = _apply_cli(device, command, ctx, _fabric_model)
    _cli_contexts[device_id] = new_ctx
    _fabric_configs.clear()

    # Update link references if hostname changed
    if device.hostname != old_hostname:
        for link in _fabric_model.links:
            if link.from_device == old_hostname:
                link.from_device = device.hostname
            if link.to_device == old_hostname:
                link.to_device = device.hostname

    prompt = _build_prompt(device, new_ctx)
    return {"device": device.hostname, "result": result, "model": device.to_dict(), "prompt": prompt}


_cli_contexts: dict = {}

# NX-OS command and keyword abbreviation expansion
_NXOS_ABBREVIATIONS = {
    "int": "interface",
    "intf": "interface",
    "eth": "Ethernet",
    "po": "port-channel",
    "lo": "loopback",
    "mgmt": "mgmt",
    "vl": "Vlan",
    "desc": "description",
    "shut": "shutdown",
    "sw": "switchport",
    "ro": "router",
    "nei": "neighbor",
    "nbr": "neighbor",
    "addr": "address",
    "add": "address",
    "fam": "family",
    "uni": "unicast",
    "multi": "multicast",
    "conf": "config",
    "feat": "feature",
    "bgp": "bgp",
    "evp": "evpn",
    "vp": "vpc",
    "mem": "member",
    "src": "source",
    "dst": "destination",
    "pref": "prefix-list",
    "rm": "route-map",
    "redis": "redistribute",
    "dir": "direct",
    "conn": "connected",
    "stat": "static",
    "proto": "protocol",
    "host": "host-reachability",
    "ingr": "ingress-replication",
    "mcast": "mcast-group",
    "sys": "system-priority",
    "del": "delay",
    "rest": "restore",
    "auto": "auto-recovery",
    "pk": "peer-keepalive",
    "pg": "peer-gateway",
    "pl": "peer-link",
    "chn": "channel-group",
    "rem": "remote-as",
    "upd": "update-source",
    "ebgp": "ebgp-multihop",
    "sen": "send-community",
    "rew": "rewrite-evpn-rt-asn",
    "rr": "route-reflector-client",
    "all": "allowas-in",
    "dis": "disable-peer-as-check",
    "ret": "retain",
    "rt": "route-target",
    "rd": "rd",
    "vni": "vni",
    "fab": "fabric",
    "fwd": "forwarding",
    "any": "anycast-gateway",
    "run": "running-config",
    "ver": "version",
    "nve": "nve1",
}

_INTF_EXPANSIONS = {
    "eth": "Ethernet",
    "e": "Ethernet",
    "ethernet": "Ethernet",
    "po": "port-channel",
    "port-channel": "port-channel",
    "lo": "loopback",
    "loopback": "loopback",
    "vlan": "Vlan",
    "vl": "Vlan",
    "mgmt": "mgmt",
    "nve": "nve",
}


def _expand_command(command: str) -> str:
    """Expand NX-OS abbreviations to full commands."""
    parts = command.split()
    if not parts:
        return command

    expanded = []
    i = 0
    while i < len(parts):
        token = parts[i]
        token_lower = token.lower()

        # First token: expand command keywords
        if i == 0:
            if token_lower in ("int", "intf"):
                expanded.append("interface")
            elif token_lower in ("no",):
                expanded.append("no")
            elif token_lower in ("sh", "sho"):
                expanded.append("show")
            elif token_lower in ("ro", "rout"):
                expanded.append("router")
            elif token_lower in ("feat",):
                expanded.append("feature")
            elif token_lower in ("desc",):
                expanded.append("description")
            elif token_lower in ("sw",):
                expanded.append("switchport")
            elif token_lower in ("chn", "chan"):
                expanded.append("channel-group")
            elif token_lower in ("nei", "nbr"):
                expanded.append("neighbor")
            elif token_lower in ("addr-fam", "address-fam"):
                expanded.append("address-family")
            elif token_lower in ("addr", "add") and i + 1 < len(parts) and parts[i+1].lower().startswith("fam"):
                expanded.append("address-family")
                i += 1
            else:
                expanded.append(token)
        # If "no" was first, the next token should also get command-level expansion
        elif expanded and expanded[0].lower() == "no" and i == 1:
            if token_lower in ("int", "intf"):
                expanded.append("interface")
            elif token_lower in ("shut",):
                expanded.append("shutdown")
            elif token_lower in ("sw",):
                expanded.append("switchport")
            elif token_lower in ("feat",):
                expanded.append("feature")
            elif token_lower in _NXOS_ABBREVIATIONS:
                expanded.append(_NXOS_ABBREVIATIONS[token_lower])
            else:
                expanded.append(token)
        # Interface name expansion: only when `interface` is the command or `no interface`
        elif expanded and (
            (expanded[0].lower() == "interface" and i == 1) or
            (expanded[0].lower() == "no" and len(expanded) > 1 and expanded[1].lower() == "interface" and i == 2)
        ):
            expanded_name = _expand_interface_name(parts[i:])
            expanded.append(expanded_name)
            break
        elif expanded and expanded[0].lower() == "interface" and token_lower == "nve1":
            expanded.append("nve1")
        else:
            # General abbreviation expansion for remaining tokens
            if token_lower in _NXOS_ABBREVIATIONS:
                expanded.append(_NXOS_ABBREVIATIONS[token_lower])
            else:
                expanded.append(token)
        i += 1

    return " ".join(expanded)


def _expand_interface_name(tokens: list) -> str:
    """Expand interface name tokens like 'eth 1/60' or 'eth1/60' to 'Ethernet1/60'."""
    if not tokens:
        return ""

    first = tokens[0]
    first_lower = first.lower()

    # Check if the first token contains a prefix and port combined (e.g. "eth1/60")
    for prefix, full in _INTF_EXPANSIONS.items():
        if first_lower.startswith(prefix) and len(first_lower) > len(prefix):
            remainder = first[len(prefix):]
            if remainder[0].isdigit() or remainder[0] == '/':
                return full + remainder

    # Check if it's a standalone prefix followed by a separate port token
    if first_lower in _INTF_EXPANSIONS:
        full_prefix = _INTF_EXPANSIONS[first_lower]
        if len(tokens) > 1:
            return full_prefix + tokens[1]
        return full_prefix

    # No expansion needed, join everything
    return " ".join(tokens)


def _build_prompt(device, ctx: dict) -> str:
    """Build the NX-OS-style prompt based on current context."""
    mode = ctx.get("mode", "config")
    sub = ctx.get("sub")
    target = ctx.get("target", "")

    if mode == "config":
        return f"{device.hostname}(config)# "
    elif mode == "interface":
        return f"{device.hostname}(config-if:{target})# "
    elif mode == "router_bgp":
        if sub == "neighbor_af":
            return f"{device.hostname}(config-router-neighbor-af)# "
        elif sub == "neighbor":
            return f"{device.hostname}(config-router-neighbor)# "
        elif sub == "vrf":
            return f"{device.hostname}(config-router-vrf)# "
        elif sub == "af":
            return f"{device.hostname}(config-router-af)# "
        return f"{device.hostname}(config-router)# "
    elif mode == "vrf":
        if sub == "af":
            return f"{device.hostname}(config-vrf-af)# "
        return f"{device.hostname}(config-vrf)# "
    elif mode == "nve":
        if sub == "member_vni":
            return f"{device.hostname}(config-if-nve-vni)# "
        return f"{device.hostname}(config-if-nve)# "
    elif mode == "evpn":
        return f"{device.hostname}(config-evpn)# "
    elif mode == "vpc":
        return f"{device.hostname}(config-vpc-domain)# "
    return f"{device.hostname}(config)# "


def _apply_cli(device, command: str, ctx: dict, model) -> tuple[str, dict]:
    """Context-aware NX-OS CLI command parser."""
    command = _expand_command(command)
    parts = command.split()
    if not parts:
        return "Empty command", ctx

    cmd = parts[0].lower()
    mode = ctx.get("mode", "config")

    # --- Navigation commands (work in all modes) ---
    if cmd == "exit":
        if mode == "config":
            return "Already at config level", ctx
        parent = _exit_context(ctx)
        return f"Exited to {parent.get('mode', 'config')} mode", parent

    if cmd == "end":
        return "Returned to config mode", {"mode": "config", "sub": None, "target": None}

    # --- Show commands (work in all modes) ---
    if cmd == "show":
        return _handle_show(device, parts, model), ctx

    if cmd in ("help", "?"):
        return _handle_help(ctx), ctx

    # --- Dispatch based on current mode ---
    if mode == "config":
        return _apply_config_mode(device, parts, ctx, model)
    elif mode == "interface":
        return _apply_interface_mode(device, parts, ctx)
    elif mode == "router_bgp":
        return _apply_router_bgp_mode(device, parts, ctx, model)
    elif mode == "vrf":
        return _apply_vrf_mode(device, parts, ctx, model)
    elif mode == "nve":
        return _apply_nve_mode(device, parts, ctx, model)
    elif mode == "evpn":
        return _apply_evpn_mode(device, parts, ctx, model)
    elif mode == "vpc":
        return _apply_vpc_mode(device, parts, ctx)

    return f"Command applied: {command}", ctx


def _exit_context(ctx: dict) -> dict:
    """Move up one context level."""
    mode = ctx.get("mode", "config")
    sub = ctx.get("sub")

    if mode in ("interface", "vrf", "nve", "evpn", "vpc"):
        if sub:
            return {"mode": mode, "sub": None, "target": ctx.get("target")}
        return {"mode": "config", "sub": None, "target": None}
    if mode == "router_bgp":
        if sub == "neighbor_af":
            return {"mode": "router_bgp", "sub": "neighbor", "target": ctx.get("target"), "neighbor": ctx.get("neighbor")}
        if sub in ("neighbor", "vrf", "af"):
            return {"mode": "router_bgp", "sub": None, "target": ctx.get("target")}
        return {"mode": "config", "sub": None, "target": None}
    return {"mode": "config", "sub": None, "target": None}


# =============================================================================
# CONFIG MODE (top level)
# =============================================================================

def _apply_config_mode(device, parts: list, ctx: dict, model) -> tuple[str, dict]:
    """Handle commands at the config terminal level."""
    cmd = parts[0].lower()

    if cmd == "hostname" and len(parts) > 1:
        device.hostname = parts[1]
        return f"Hostname set to {parts[1]}", ctx

    if cmd == "interface" and len(parts) > 1:
        intf_name = " ".join(parts[1:])
        existing = next((i for i in device.interfaces if i["name"].lower() == intf_name.lower()), None)
        if not existing:
            device.interfaces.append({
                "name": intf_name, "description": "", "ip": "", "speed": "",
                "shutdown": False, "mode": "", "vlan": "", "channel_group": ""
            })
        new_ctx = {"mode": "interface", "sub": None, "target": intf_name}
        return f"Entered interface {intf_name}", new_ctx

    if cmd == "router" and len(parts) >= 3 and parts[1].lower() == "bgp":
        device.asn = parts[2]
        new_ctx = {"mode": "router_bgp", "sub": None, "target": parts[2]}
        return f"Entered router bgp {parts[2]}", new_ctx

    if cmd == "vrf" and len(parts) >= 3 and parts[1].lower() == "context":
        vrf_name = parts[2]
        device.config.setdefault("vrfs", {}).setdefault(vrf_name, {"vni": "", "rd": "auto", "rt_import": "auto", "rt_export": "auto"})
        new_ctx = {"mode": "vrf", "sub": None, "target": vrf_name}
        return f"Entered vrf context {vrf_name}", new_ctx

    if cmd == "interface" and len(parts) > 1 and parts[1].lower() == "nve1":
        new_ctx = {"mode": "nve", "sub": None, "target": "nve1"}
        return "Entered interface nve1", new_ctx

    if cmd == "evpn":
        new_ctx = {"mode": "evpn", "sub": None, "target": "evpn"}
        return "Entered evpn configuration", new_ctx

    if cmd == "vpc" and len(parts) >= 3 and parts[1].lower() == "domain":
        device.vpc_domain = parts[2]
        new_ctx = {"mode": "vpc", "sub": None, "target": parts[2]}
        return f"Entered vpc domain {parts[2]}", new_ctx

    if cmd == "role" and len(parts) > 1:
        valid_roles = ("spine", "leaf", "border_leaf", "border_gateway", "super_spine", "service_leaf")
        new_role = parts[1].lower()
        if new_role in valid_roles:
            device.role = new_role
            return f"Role changed to {new_role}", ctx
        return f"Invalid role. Valid: {', '.join(valid_roles)}", ctx

    if cmd == "loopback0" and len(parts) > 1:
        device.loopback0 = parts[1] if "/" in parts[1] else parts[1] + "/32"
        return f"Loopback0 set to {device.loopback0}", ctx

    if cmd == "loopback1" and len(parts) > 1:
        device.loopback1 = parts[1] if "/" in parts[1] else parts[1] + "/32"
        return f"Loopback1 (VTEP) set to {device.loopback1}", ctx

    if cmd == "loopback2" and len(parts) > 1:
        device.loopback2 = parts[1] if "/" in parts[1] else parts[1] + "/32"
        return f"Loopback2 (Multi-site) set to {device.loopback2}", ctx

    if cmd == "site" and len(parts) > 1:
        device.site = parts[1]
        return f"Site set to {parts[1]}", ctx

    if cmd == "mgmt-ip" and len(parts) > 1:
        device.mgmt_ip = parts[1]
        return f"Management IP set to {parts[1]}", ctx

    if cmd == "ip" and len(parts) >= 3 and parts[1].lower() == "address":
        device.config.setdefault("ip_addresses", []).append(" ".join(parts[2:]))
        return f"IP address added: {' '.join(parts[2:])}", ctx

    if cmd == "ip" and len(parts) >= 3 and parts[1].lower() == "route":
        device.config.setdefault("static_routes", []).append(" ".join(parts[2:]))
        return f"Static route added: {' '.join(parts[2:])}", ctx

    if cmd == "ip" and len(parts) >= 3 and parts[1].lower() == "prefix-list":
        device.config.setdefault("prefix_lists", []).append(" ".join(parts[2:]))
        return f"Prefix-list added: {' '.join(parts[2:])}", ctx

    if cmd == "route-map" and len(parts) >= 2:
        device.config.setdefault("route_maps", []).append(" ".join(parts[1:]))
        return f"Route-map added: {' '.join(parts[1:])}", ctx

    if cmd == "ntp" and len(parts) >= 3 and parts[1].lower() == "server":
        device.config.setdefault("ntp_servers", []).append(parts[2])
        return f"NTP server added: {parts[2]}", ctx

    if cmd == "feature" and len(parts) > 1:
        device.config.setdefault("features", []).append(parts[1])
        return f"Feature enabled: {parts[1]}", ctx

    if cmd == "no" and len(parts) > 1:
        return _handle_no_cmd(device, parts[1:], ctx)

    return f"Command applied: {' '.join(parts)}", ctx


# =============================================================================
# INTERFACE MODE
# =============================================================================

def _apply_interface_mode(device, parts: list, ctx: dict) -> tuple[str, dict]:
    """Handle commands within interface context."""
    cmd = parts[0].lower()
    intf_name = ctx.get("target", "")
    intf = next((i for i in device.interfaces if i["name"].lower() == intf_name.lower()), None)
    if not intf:
        return f"Interface {intf_name} not found", ctx

    if cmd == "description" and len(parts) > 1:
        intf["description"] = " ".join(parts[1:])
        return f"Description: {intf['description']}", ctx

    if cmd == "ip" and len(parts) >= 3 and parts[1].lower() == "address":
        intf["ip"] = " ".join(parts[2:])
        return f"IP address: {intf['ip']}", ctx

    if cmd == "speed" and len(parts) > 1:
        intf["speed"] = parts[1]
        return f"Speed set to {parts[1]}", ctx

    if cmd == "switchport" and len(parts) >= 3:
        if parts[1].lower() == "mode":
            intf["mode"] = parts[2].lower()
            return f"Switchport mode: {parts[2]}", ctx
        if parts[1].lower() == "access" and len(parts) >= 4 and parts[2].lower() == "vlan":
            intf["vlan"] = parts[3]
            return f"Access VLAN: {parts[3]}", ctx
        if parts[1].lower() == "trunk" and len(parts) >= 5 and parts[2].lower() == "allowed":
            intf.setdefault("trunk_vlans", parts[4])
            return f"Trunk allowed VLANs: {parts[4]}", ctx

    if cmd == "channel-group" and len(parts) >= 2:
        intf["channel_group"] = parts[1]
        mode = parts[3] if len(parts) >= 4 and parts[2].lower() == "mode" else ""
        if mode:
            intf["channel_mode"] = mode
        return f"Channel-group {parts[1]}" + (f" mode {mode}" if mode else ""), ctx

    if cmd == "shutdown":
        intf["shutdown"] = True
        return "Interface shutdown", ctx

    if cmd == "no" and len(parts) > 1:
        subcmd = parts[1].lower()
        if subcmd == "shutdown":
            intf["shutdown"] = False
            return "Interface no shutdown", ctx
        if subcmd == "switchport":
            intf["mode"] = "routed"
            return "Switchport disabled (routed mode)", ctx

    if cmd == "mtu" and len(parts) > 1:
        intf["mtu"] = parts[1]
        return f"MTU set to {parts[1]}", ctx

    if cmd == "ip" and len(parts) >= 2 and parts[1].lower() == "forward":
        intf["ip_forward"] = True
        return "IP forward enabled", ctx

    if cmd == "fabric" and " ".join(parts[1:4]).lower() == "forwarding mode anycast-gateway":
        intf["anycast_gw"] = True
        return "Anycast gateway mode enabled", ctx

    if cmd == "vrf" and len(parts) >= 3 and parts[1].lower() == "member":
        intf["vrf"] = parts[2]
        return f"VRF member: {parts[2]}", ctx

    return f"({intf_name}) Command applied: {' '.join(parts)}", ctx


# =============================================================================
# ROUTER BGP MODE
# =============================================================================

def _apply_router_bgp_mode(device, parts: list, ctx: dict, model) -> tuple[str, dict]:
    """Handle commands within router bgp context."""
    cmd = parts[0].lower()
    sub = ctx.get("sub")
    bgp = device.config.setdefault("bgp", {"router_id": "", "neighbors": {}, "vrfs": {}, "address_families": {}})

    if cmd == "router-id" and len(parts) > 1:
        bgp["router_id"] = parts[1]
        return f"Router-ID: {parts[1]}", ctx

    if cmd == "address-family" and len(parts) >= 3:
        af_name = " ".join(parts[1:])
        bgp.setdefault("address_families", {})[af_name] = bgp.get("address_families", {}).get(af_name, {})
        new_ctx = dict(ctx)
        new_ctx["sub"] = "af"
        new_ctx["af"] = af_name
        return f"Entered address-family {af_name}", new_ctx

    if cmd == "neighbor" and len(parts) >= 2:
        nbr_ip = parts[1]
        bgp.setdefault("neighbors", {})[nbr_ip] = bgp.get("neighbors", {}).get(nbr_ip, {})
        new_ctx = dict(ctx)
        new_ctx["sub"] = "neighbor"
        new_ctx["neighbor"] = nbr_ip
        if len(parts) >= 3:
            return _apply_bgp_neighbor_inline(bgp["neighbors"][nbr_ip], parts[2:], new_ctx)
        return f"Entered neighbor {nbr_ip}", new_ctx

    if cmd == "vrf" and len(parts) >= 2:
        vrf_name = parts[1]
        bgp.setdefault("vrfs", {})[vrf_name] = bgp.get("vrfs", {}).get(vrf_name, {})
        new_ctx = dict(ctx)
        new_ctx["sub"] = "vrf"
        new_ctx["vrf"] = vrf_name
        return f"Entered BGP vrf {vrf_name}", new_ctx

    if cmd == "log-neighbor-changes":
        bgp["log_neighbor_changes"] = True
        return "Log neighbor changes enabled", ctx

    # Sub-context: neighbor
    if sub == "neighbor":
        nbr_ip = ctx.get("neighbor", "")
        nbr = bgp.get("neighbors", {}).get(nbr_ip, {})
        return _apply_bgp_neighbor_cmd(nbr, parts, ctx)

    # Sub-context: address-family
    if sub == "af":
        af_name = ctx.get("af", "")
        af_cfg = bgp.get("address_families", {}).get(af_name, {})
        return _apply_bgp_af_cmd(af_cfg, parts, ctx)

    # Sub-context: neighbor address-family
    if sub == "neighbor_af":
        nbr_ip = ctx.get("neighbor", "")
        nbr = bgp.get("neighbors", {}).get(nbr_ip, {})
        af_name = ctx.get("af", "")
        af_cfg = nbr.setdefault("address_families", {}).setdefault(af_name, {})
        return _apply_bgp_neighbor_af_cmd(af_cfg, parts, ctx)

    # Sub-context: vrf
    if sub == "vrf":
        vrf_name = ctx.get("vrf", "")
        vrf_cfg = bgp.get("vrfs", {}).get(vrf_name, {})
        return _apply_bgp_vrf_cmd(vrf_cfg, parts, ctx)

    return f"(router-bgp) Command applied: {' '.join(parts)}", ctx


def _apply_bgp_neighbor_inline(nbr: dict, parts: list, ctx: dict) -> tuple[str, dict]:
    """Handle inline neighbor sub-commands like 'neighbor x.x.x.x remote-as 65000'."""
    return _apply_bgp_neighbor_cmd(nbr, parts, ctx)


def _apply_bgp_neighbor_cmd(nbr: dict, parts: list, ctx: dict) -> tuple[str, dict]:
    """Commands within BGP neighbor context."""
    cmd = parts[0].lower()

    if cmd == "remote-as" and len(parts) > 1:
        nbr["remote_as"] = parts[1]
        return f"Remote-AS: {parts[1]}", ctx

    if cmd == "update-source" and len(parts) > 1:
        nbr["update_source"] = parts[1]
        return f"Update-source: {parts[1]}", ctx

    if cmd == "ebgp-multihop" and len(parts) > 1:
        nbr["ebgp_multihop"] = parts[1]
        return f"eBGP multihop: {parts[1]}", ctx

    if cmd == "peer-type" and len(parts) > 1:
        nbr["peer_type"] = " ".join(parts[1:])
        return f"Peer-type: {' '.join(parts[1:])}", ctx

    if cmd == "description" and len(parts) > 1:
        nbr["description"] = " ".join(parts[1:])
        return f"Description: {' '.join(parts[1:])}", ctx

    if cmd == "address-family" and len(parts) >= 3:
        af_name = " ".join(parts[1:])
        nbr.setdefault("address_families", {})[af_name] = nbr.get("address_families", {}).get(af_name, {})
        new_ctx = dict(ctx)
        new_ctx["sub"] = "neighbor_af"
        new_ctx["af"] = af_name
        return f"Entered neighbor address-family {af_name}", new_ctx

    if cmd == "send-community" and len(parts) >= 1:
        val = parts[1] if len(parts) > 1 else "both"
        nbr["send_community"] = val
        return f"Send-community: {val}", ctx

    if cmd == "allowas-in" and len(parts) >= 1:
        val = parts[1] if len(parts) > 1 else "3"
        nbr["allowas_in"] = val
        return f"Allowas-in: {val}", ctx

    if cmd == "disable-peer-as-check":
        nbr["disable_peer_as_check"] = True
        return "Disable-peer-as-check enabled", ctx

    return f"(neighbor) Command applied: {' '.join(parts)}", ctx


def _apply_bgp_neighbor_af_cmd(af_cfg: dict, parts: list, ctx: dict) -> tuple[str, dict]:
    """Commands within BGP neighbor address-family context."""
    cmd = parts[0].lower()

    if cmd == "send-community":
        val = parts[1] if len(parts) > 1 else "both"
        af_cfg["send_community"] = val
        return f"Send-community: {val}", ctx

    if cmd == "rewrite-evpn-rt-asn":
        af_cfg["rewrite_evpn_rt_asn"] = True
        return "Rewrite-evpn-rt-asn enabled", ctx

    if cmd == "route-reflector-client":
        af_cfg["route_reflector_client"] = True
        return "Route-reflector-client enabled", ctx

    if cmd == "allowas-in" and len(parts) >= 1:
        af_cfg["allowas_in"] = parts[1] if len(parts) > 1 else "3"
        return f"Allowas-in: {af_cfg['allowas_in']}", ctx

    if cmd == "disable-peer-as-check":
        af_cfg["disable_peer_as_check"] = True
        return "Disable-peer-as-check enabled", ctx

    return f"(neighbor-af) Command applied: {' '.join(parts)}", ctx


def _apply_bgp_af_cmd(af_cfg: dict, parts: list, ctx: dict) -> tuple[str, dict]:
    """Commands within BGP address-family context."""
    cmd = parts[0].lower()

    if cmd == "network" and len(parts) > 1:
        af_cfg.setdefault("networks", []).append(" ".join(parts[1:]))
        return f"Network: {' '.join(parts[1:])}", ctx

    if cmd == "retain" and " ".join(parts[1:3]).lower() == "route-target all":
        af_cfg["retain_rt_all"] = True
        return "Retain route-target all", ctx

    if cmd == "advertise" and len(parts) > 1:
        af_cfg.setdefault("advertise", []).append(" ".join(parts[1:]))
        return f"Advertise: {' '.join(parts[1:])}", ctx

    if cmd == "redistribute" and len(parts) > 1:
        af_cfg.setdefault("redistribute", []).append(" ".join(parts[1:]))
        return f"Redistribute: {' '.join(parts[1:])}", ctx

    if cmd == "maximum-paths" and len(parts) > 1:
        af_cfg["max_paths"] = parts[1]
        return f"Maximum-paths: {parts[1]}", ctx

    return f"(address-family) Command applied: {' '.join(parts)}", ctx


def _apply_bgp_vrf_cmd(vrf_cfg: dict, parts: list, ctx: dict) -> tuple[str, dict]:
    """Commands within BGP vrf context."""
    cmd = parts[0].lower()

    if cmd == "address-family" and len(parts) >= 3:
        af_name = " ".join(parts[1:])
        vrf_cfg.setdefault("address_families", {})[af_name] = vrf_cfg.get("address_families", {}).get(af_name, {})
        new_ctx = dict(ctx)
        new_ctx["sub"] = "af"
        new_ctx["af"] = af_name
        return f"Entered BGP VRF address-family {af_name}", new_ctx

    if cmd == "redistribute" and len(parts) > 1:
        vrf_cfg.setdefault("redistribute", []).append(" ".join(parts[1:]))
        return f"Redistribute: {' '.join(parts[1:])}", ctx

    return f"(bgp-vrf) Command applied: {' '.join(parts)}", ctx


# =============================================================================
# VRF CONTEXT MODE
# =============================================================================

def _apply_vrf_mode(device, parts: list, ctx: dict, model) -> tuple[str, dict]:
    """Handle commands within vrf context."""
    cmd = parts[0].lower()
    vrf_name = ctx.get("target", "")
    vrf_cfg = device.config.get("vrfs", {}).get(vrf_name, {})

    if cmd == "vni" and len(parts) > 1:
        vrf_cfg["vni"] = parts[1]
        return f"VNI: {parts[1]}", ctx

    if cmd == "rd" and len(parts) > 1:
        vrf_cfg["rd"] = parts[1]
        return f"RD: {parts[1]}", ctx

    if cmd == "address-family" and len(parts) >= 3:
        af_name = " ".join(parts[1:])
        vrf_cfg.setdefault("address_families", {})[af_name] = {}
        new_ctx = dict(ctx)
        new_ctx["sub"] = "af"
        new_ctx["af"] = af_name
        return f"Entered VRF address-family {af_name}", new_ctx

    if ctx.get("sub") == "af":
        af_name = ctx.get("af", "")
        af_cfg = vrf_cfg.get("address_families", {}).get(af_name, {})
        if cmd == "route-target" and len(parts) >= 3:
            direction = parts[1].lower()
            value = " ".join(parts[2:])
            af_cfg.setdefault("route_targets", []).append({"direction": direction, "value": value})
            return f"Route-target {direction}: {value}", ctx
        if cmd == "redistribute" and len(parts) > 1:
            af_cfg.setdefault("redistribute", []).append(" ".join(parts[1:]))
            return f"Redistribute: {' '.join(parts[1:])}", ctx

    return f"(vrf) Command applied: {' '.join(parts)}", ctx


# =============================================================================
# NVE MODE
# =============================================================================

def _apply_nve_mode(device, parts: list, ctx: dict, model) -> tuple[str, dict]:
    """Handle commands within interface nve1 context."""
    cmd = parts[0].lower()
    nve = device.config.setdefault("nve", {"source_interface": "", "members": {}, "multisite_bgw_intf": ""})

    if cmd == "source-interface" and len(parts) > 1:
        nve["source_interface"] = parts[1]
        return f"Source-interface: {parts[1]}", ctx

    if cmd == "host-reachability" and len(parts) >= 3:
        nve["host_reachability"] = parts[2]
        return f"Host-reachability protocol: {parts[2]}", ctx

    if cmd == "multisite" and len(parts) >= 4 and parts[1].lower() == "border-gateway":
        nve["multisite_bgw_intf"] = " ".join(parts[3:])
        return f"Multisite border-gateway interface: {' '.join(parts[3:])}", ctx

    if cmd == "member" and len(parts) >= 3 and parts[1].lower() == "vni":
        vni = parts[2]
        associate_vrf = "associate-vrf" in " ".join(parts[3:]).lower()
        nve.setdefault("members", {})[vni] = nve.get("members", {}).get(vni, {})
        nve["members"][vni]["associate_vrf"] = associate_vrf
        new_ctx = dict(ctx)
        new_ctx["sub"] = "member_vni"
        new_ctx["vni"] = vni
        return f"Entered member vni {vni}" + (" associate-vrf" if associate_vrf else ""), new_ctx

    if ctx.get("sub") == "member_vni":
        vni = ctx.get("vni", "")
        vni_cfg = nve.get("members", {}).get(vni, {})
        if cmd == "multisite" and len(parts) >= 2 and parts[1].lower() == "ingress-replication":
            vni_cfg["multisite_ir"] = True
            return "Multisite ingress-replication enabled", ctx
        if cmd == "ingress-replication" and len(parts) >= 3:
            vni_cfg["ingress_replication"] = parts[2]
            return f"Ingress-replication protocol: {parts[2]}", ctx
        if cmd == "mcast-group" and len(parts) > 1:
            vni_cfg["mcast_group"] = parts[1]
            return f"Mcast-group: {parts[1]}", ctx

    if cmd == "no" and len(parts) > 1 and parts[1].lower() == "shutdown":
        nve["shutdown"] = False
        return "NVE no shutdown", ctx

    if cmd == "shutdown":
        nve["shutdown"] = True
        return "NVE shutdown", ctx

    return f"(nve) Command applied: {' '.join(parts)}", ctx


# =============================================================================
# EVPN MODE
# =============================================================================

def _apply_evpn_mode(device, parts: list, ctx: dict, model) -> tuple[str, dict]:
    """Handle commands within evpn context."""
    cmd = parts[0].lower()
    evpn = device.config.setdefault("evpn", {"vnis": {}, "multisite": {}})

    if cmd == "vni" and len(parts) >= 3:
        vni = parts[1]
        vni_type = parts[2] if len(parts) > 2 else "l2"
        evpn.setdefault("vnis", {})[vni] = evpn.get("vnis", {}).get(vni, {"type": vni_type})
        return f"EVPN VNI {vni} {vni_type} configured", ctx

    if cmd == "multisite" and len(parts) >= 3 and parts[1].lower() == "border-gateway":
        evpn["multisite"]["border_gateway_id"] = parts[2]
        return f"Multisite border-gateway: {parts[2]}", ctx

    if cmd == "rd" and len(parts) > 1:
        evpn["rd"] = parts[1]
        return f"RD: {parts[1]}", ctx

    if cmd == "route-target" and len(parts) >= 3:
        direction = parts[1]
        value = parts[2]
        evpn.setdefault("route_targets", []).append({"direction": direction, "value": value})
        return f"Route-target {direction}: {value}", ctx

    return f"(evpn) Command applied: {' '.join(parts)}", ctx


# =============================================================================
# VPC DOMAIN MODE
# =============================================================================

def _apply_vpc_mode(device, parts: list, ctx: dict) -> tuple[str, dict]:
    """Handle commands within vpc domain context."""
    cmd = parts[0].lower()
    vpc = device.config.setdefault("vpc", {"domain": ctx.get("target", ""), "peer_keepalive": "", "peer_link": ""})

    if cmd == "peer-keepalive" and len(parts) >= 3:
        vpc["peer_keepalive"] = " ".join(parts[1:])
        return f"Peer-keepalive: {' '.join(parts[1:])}", ctx

    if cmd == "peer-link" and len(parts) > 1:
        vpc["peer_link"] = parts[1]
        return f"Peer-link: {parts[1]}", ctx

    if cmd == "role" and len(parts) > 1:
        vpc["role_priority"] = parts[1]
        return f"Role priority: {parts[1]}", ctx

    if cmd == "system-priority" and len(parts) > 1:
        vpc["system_priority"] = parts[1]
        return f"System-priority: {parts[1]}", ctx

    if cmd == "auto-recovery":
        vpc["auto_recovery"] = True
        return "Auto-recovery enabled", ctx

    if cmd == "delay" and len(parts) >= 4 and parts[1].lower() == "restore":
        vpc["delay_restore"] = parts[2]
        return f"Delay restore: {parts[2]}", ctx

    if cmd == "peer-gateway":
        vpc["peer_gateway"] = True
        return "Peer-gateway enabled", ctx

    if cmd == "ip" and len(parts) >= 4 and parts[1].lower() == "arp" and parts[2].lower() == "synchronize":
        vpc["arp_sync"] = True
        return "IP ARP synchronize enabled", ctx

    return f"(vpc-domain) Command applied: {' '.join(parts)}", ctx


# =============================================================================
# HELPERS
# =============================================================================

def _handle_no_cmd(device, parts: list, ctx: dict) -> tuple[str, dict]:
    """Handle 'no' negation commands at config level."""
    cmd = parts[0].lower()
    if cmd == "interface" and len(parts) > 1:
        intf_name = " ".join(parts[1:])
        device.interfaces = [i for i in device.interfaces if i["name"].lower() != intf_name.lower()]
        return f"Interface {intf_name} removed", ctx
    if cmd == "feature" and len(parts) > 1:
        feats = device.config.get("features", [])
        device.config["features"] = [f for f in feats if f != parts[1]]
        return f"Feature {parts[1]} disabled", ctx
    return f"no {' '.join(parts)} applied", ctx


def _handle_show(device, parts: list, model) -> str:
    """Handle show commands."""
    if len(parts) < 2:
        return "Incomplete show command"
    subcmd = parts[1].lower()

    if subcmd in ("running-config", "run"):
        lines = [f"! Device: {device.hostname}", f"hostname {device.hostname}"]
        lines.append(f"!")
        if device.config.get("features"):
            for feat in device.config["features"]:
                lines.append(f"feature {feat}")
            lines.append("!")
        lines.append(f"interface loopback0\n  ip address {device.loopback0 or 'not configured'}")
        if device.loopback1:
            lines.append(f"interface loopback1\n  ip address {device.loopback1}")
        if device.loopback2:
            lines.append(f"interface loopback2\n  ip address {device.loopback2}")
        lines.append(f"!")
        if device.vpc_domain:
            lines.append(f"vpc domain {device.vpc_domain}")
            if device.vpc_peer:
                lines.append(f"  peer {device.vpc_peer}")
        lines.append(f"!")
        for intf in device.interfaces:
            lines.append(f"interface {intf['name']}")
            if intf.get("description"):
                lines.append(f"  description {intf['description']}")
            if intf.get("ip"):
                lines.append(f"  ip address {intf['ip']}")
            if intf.get("vrf"):
                lines.append(f"  vrf member {intf['vrf']}")
            if intf.get("shutdown"):
                lines.append(f"  shutdown")
            else:
                lines.append(f"  no shutdown")
        lines.append(f"!")
        if device.asn:
            lines.append(f"router bgp {device.asn}")
            lines.append(f"  router-id {device.loopback0.split('/')[0] if device.loopback0 else '0.0.0.0'}")
            bgp = device.config.get("bgp", {})
            for nbr_ip, nbr_cfg in bgp.get("neighbors", {}).items():
                lines.append(f"  neighbor {nbr_ip}")
                if nbr_cfg.get("remote_as"):
                    lines.append(f"    remote-as {nbr_cfg['remote_as']}")
                if nbr_cfg.get("update_source"):
                    lines.append(f"    update-source {nbr_cfg['update_source']}")
                if nbr_cfg.get("peer_type"):
                    lines.append(f"    peer-type {nbr_cfg['peer_type']}")
                for af_name, af_cfg in nbr_cfg.get("address_families", {}).items():
                    lines.append(f"    address-family {af_name}")
                    if af_cfg.get("send_community"):
                        lines.append(f"      send-community {af_cfg['send_community']}")
                    if af_cfg.get("rewrite_evpn_rt_asn"):
                        lines.append(f"      rewrite-evpn-rt-asn")
        lines.append(f"!")
        for vrf_name, vrf_cfg in device.config.get("vrfs", {}).items():
            lines.append(f"vrf context {vrf_name}")
            if vrf_cfg.get("vni"):
                lines.append(f"  vni {vrf_cfg['vni']}")
            if vrf_cfg.get("rd"):
                lines.append(f"  rd {vrf_cfg['rd']}")
        return "\n".join(lines)

    if subcmd in ("interfaces", "interface"):
        if not device.interfaces:
            return "No interfaces configured"
        lines = [f"{'Interface':<25} {'Status':<10} {'Description'}"]
        lines.append("-" * 60)
        for intf in device.interfaces:
            status = "down" if intf.get("shutdown") else "up"
            lines.append(f"{intf['name']:<25} {status:<10} {intf.get('description', '')}")
        return "\n".join(lines)

    if subcmd == "version":
        return f"{device.hostname} | Model: {device.model or 'N9K'} | Role: {device.role} | Site: {device.site} | ASN: {device.asn or 'N/A'}"

    if subcmd == "bgp" or (subcmd == "ip" and len(parts) >= 3 and parts[2].lower() == "bgp"):
        bgp = device.config.get("bgp", {})
        lines = [f"BGP ASN: {device.asn}", f"Router-ID: {bgp.get('router_id', device.loopback0)}"]
        for nbr_ip, nbr_cfg in bgp.get("neighbors", {}).items():
            lines.append(f"  Neighbor {nbr_ip} remote-as {nbr_cfg.get('remote_as', '?')}")
        return "\n".join(lines) if lines else "No BGP configuration"

    if subcmd == "vrf":
        vrfs = device.config.get("vrfs", {})
        if not vrfs:
            return "No VRFs configured"
        lines = []
        for name, cfg in vrfs.items():
            lines.append(f"  VRF {name}: VNI={cfg.get('vni', 'N/A')} RD={cfg.get('rd', 'auto')}")
        return "\n".join(lines)

    if subcmd == "nve":
        nve = device.config.get("nve", {})
        if not nve:
            return "NVE not configured"
        lines = [f"Source: {nve.get('source_interface', 'N/A')}"]
        for vni, cfg in nve.get("members", {}).items():
            lines.append(f"  VNI {vni}: {'associate-vrf' if cfg.get('associate_vrf') else 'L2'}")
        return "\n".join(lines)

    return f"Unknown show command: {subcmd}"


def _handle_help(ctx: dict) -> str:
    """Context-aware help."""
    mode = ctx.get("mode", "config")

    if mode == "config":
        return ("Config mode commands:\n"
                "  hostname <name>            - Set device hostname\n"
                "  interface <name>           - Enter interface config mode\n"
                "  router bgp <asn>           - Enter BGP router config\n"
                "  vrf context <name>         - Enter VRF config\n"
                "  evpn                       - Enter EVPN config\n"
                "  vpc domain <id>            - Enter vPC domain config\n"
                "  role <role>                - Set device role\n"
                "  site <name>                - Set site\n"
                "  loopback0/1/2 <ip/mask>    - Set loopback IPs\n"
                "  mgmt-ip <ip/mask>          - Set management IP\n"
                "  ip route <prefix> <nh>     - Add static route\n"
                "  ip prefix-list <...>       - Add prefix-list\n"
                "  route-map <...>            - Add route-map\n"
                "  feature <name>             - Enable feature\n"
                "  no <command>               - Negate/remove\n"
                "  show run|interfaces|bgp|vrf|nve|version\n"
                "  exit / end                 - Navigate up/to top")

    if mode == "interface":
        return ("Interface mode commands:\n"
                "  description <text>              - Set description\n"
                "  ip address <ip/mask>            - Set IP address\n"
                "  speed <value>                   - Set speed\n"
                "  mtu <value>                     - Set MTU\n"
                "  switchport mode <access|trunk>  - Set switchport mode\n"
                "  switchport access vlan <id>     - Set access VLAN\n"
                "  switchport trunk allowed vlan <list>\n"
                "  channel-group <id> mode <active|passive>\n"
                "  vrf member <name>               - Assign VRF\n"
                "  fabric forwarding mode anycast-gateway\n"
                "  ip forward                      - Enable IP forwarding\n"
                "  shutdown / no shutdown\n"
                "  exit                            - Back to config mode")

    if mode == "router_bgp":
        sub = ctx.get("sub")
        if sub == "neighbor":
            return ("BGP Neighbor commands:\n"
                    "  remote-as <asn>                 - Set remote ASN\n"
                    "  update-source <intf>            - Set update source\n"
                    "  ebgp-multihop <ttl>             - Set eBGP multihop\n"
                    "  peer-type fabric-external       - Set peer type\n"
                    "  description <text>              - Set description\n"
                    "  address-family <af>             - Enter neighbor AF\n"
                    "  send-community [both|extended]  - Send community\n"
                    "  allowas-in <count>              - Allow AS in\n"
                    "  disable-peer-as-check           - Disable AS check\n"
                    "  exit                            - Back to router-bgp")
        if sub == "neighbor_af":
            return ("BGP Neighbor Address-Family commands:\n"
                    "  send-community [both|extended]  - Send community\n"
                    "  rewrite-evpn-rt-asn             - Rewrite EVPN RT ASN\n"
                    "  route-reflector-client          - Set as RR client\n"
                    "  allowas-in <count>              - Allow AS in\n"
                    "  disable-peer-as-check           - Disable AS check\n"
                    "  exit                            - Back to neighbor")
        return ("Router BGP commands:\n"
                "  router-id <ip>                  - Set router ID\n"
                "  address-family <af>             - Enter address-family\n"
                "  neighbor <ip>                   - Enter/configure neighbor\n"
                "  vrf <name>                      - Enter BGP VRF config\n"
                "  log-neighbor-changes            - Enable logging\n"
                "  exit                            - Back to config mode")

    if mode == "vrf":
        return ("VRF Context commands:\n"
                "  vni <id>                        - Set L3 VNI\n"
                "  rd <value>                      - Set route-distinguisher\n"
                "  address-family ipv4 unicast     - Enter AF config\n"
                "    route-target <both|import|export> <value> [evpn]\n"
                "    redistribute <protocol> [route-map <name>]\n"
                "  exit                            - Back to config mode")

    if mode == "nve":
        return ("NVE Interface commands:\n"
                "  source-interface <intf>         - Set source interface\n"
                "  host-reachability protocol bgp  - Set reachability\n"
                "  multisite border-gateway interface <intf>\n"
                "  member vni <id> [associate-vrf] - Enter VNI member\n"
                "    multisite ingress-replication - Enable MS IR\n"
                "    ingress-replication protocol bgp\n"
                "    mcast-group <ip>              - Set mcast group\n"
                "  no shutdown / shutdown\n"
                "  exit                            - Back to config mode")

    if mode == "evpn":
        return ("EVPN commands:\n"
                "  vni <id> l2                     - Configure L2 VNI\n"
                "  multisite border-gateway <id>   - Set MS BGW ID\n"
                "  rd <value>                      - Set RD\n"
                "  route-target <dir> <value>      - Set RT\n"
                "  exit                            - Back to config mode")

    if mode == "vpc":
        return ("vPC Domain commands:\n"
                "  peer-keepalive destination <ip> source <ip>\n"
                "  peer-link port-channel<id>      - Set peer-link\n"
                "  role priority <value>           - Set role priority\n"
                "  system-priority <value>         - Set system priority\n"
                "  auto-recovery                   - Enable auto-recovery\n"
                "  delay restore <seconds>         - Set delay restore\n"
                "  peer-gateway                    - Enable peer-gateway\n"
                "  ip arp synchronize              - Enable ARP sync\n"
                "  exit                            - Back to config mode")

    return "Type 'help' or '?' for available commands. Use 'exit' to go up one level."


@app.get("/api/fabric/export/nxos")
async def export_nxos():
    """Download NX-OS configs as ZIP."""
    if not _fabric_model:
        raise HTTPException(status_code=404, detail="No fabric model loaded")
    exporter = NxosExporter(_fabric_model)
    zip_bytes = exporter.export()
    return Response(
        content=zip_bytes,
        media_type="application/zip",
        headers={"Content-Disposition": "attachment; filename=fabric_configs.zip"}
    )


@app.get("/api/fabric/export/yaml")
async def export_yaml():
    """Download YAML (tech-vxlan + tech-shared) as ZIP."""
    if not _fabric_model:
        raise HTTPException(status_code=404, detail="No fabric model loaded")
    exporter = YamlExporter(_fabric_model)
    yaml_files = exporter.export()

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for filename, content in yaml_files.items():
            zf.writestr(filename, content)
    zip_buffer.seek(0)

    return Response(
        content=zip_buffer.getvalue(),
        media_type="application/zip",
        headers={"Content-Disposition": "attachment; filename=fabric_yaml.zip"}
    )


@app.get("/api/fabric/export/xml")
async def export_xml():
    """Download device configs as XML (NETCONF-style payload) in a ZIP."""
    if not _fabric_model:
        raise HTTPException(status_code=404, detail="No fabric model loaded")

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for device in _fabric_model.devices:
            xml_content = _device_to_xml(device)
            zf.writestr(f"{device.hostname}.xml", xml_content)
    zip_buffer.seek(0)

    return Response(
        content=zip_buffer.getvalue(),
        media_type="application/zip",
        headers={"Content-Disposition": "attachment; filename=fabric_configs_xml.zip"}
    )


@app.get("/api/fabric/export/xml/{device_id}")
async def export_device_xml(device_id: str):
    """Download a single device config as XML."""
    if not _fabric_model:
        raise HTTPException(status_code=404, detail="No fabric model loaded")
    device = _fabric_model.get_device(device_id)
    if not device:
        raise HTTPException(status_code=404, detail="Device not found")

    xml_content = _device_to_xml(device)
    return Response(
        content=xml_content,
        media_type="application/xml",
        headers={"Content-Disposition": f"attachment; filename={device.hostname}.xml"}
    )


def _device_to_xml(device) -> str:
    """Convert a device model to NETCONF-style XML configuration."""
    from xml.etree.ElementTree import Element, SubElement, tostring
    from xml.dom.minidom import parseString

    root = Element("device-configuration", xmlns="urn:cisco:nxos:config")
    root.set("hostname", device.hostname)

    # System
    system = SubElement(root, "System")
    SubElement(system, "hostname").text = device.hostname
    SubElement(system, "role").text = device.role
    if device.model:
        SubElement(system, "model").text = device.model
    if device.site:
        SubElement(system, "site").text = device.site
    if device.mgmt_ip:
        SubElement(system, "mgmt-ip").text = device.mgmt_ip

    # Features
    features = device.config.get("features", [])
    if features:
        feat_el = SubElement(root, "features")
        for f in features:
            SubElement(feat_el, "feature").text = f

    # Loopbacks
    loopbacks = SubElement(root, "loopback-interfaces")
    if device.loopback0:
        lo0 = SubElement(loopbacks, "interface", name="loopback0")
        SubElement(lo0, "ip-address").text = device.loopback0
    if device.loopback1:
        lo1 = SubElement(loopbacks, "interface", name="loopback1")
        SubElement(lo1, "ip-address").text = device.loopback1
        SubElement(lo1, "description").text = "VTEP"
    if device.loopback2:
        lo2 = SubElement(loopbacks, "interface", name="loopback2")
        SubElement(lo2, "ip-address").text = device.loopback2
        SubElement(lo2, "description").text = "Multi-site BGW"

    # Interfaces
    if device.interfaces:
        intfs = SubElement(root, "interfaces")
        for intf in device.interfaces:
            intf_el = SubElement(intfs, "interface", name=intf["name"])
            if intf.get("description"):
                SubElement(intf_el, "description").text = intf["description"]
            if intf.get("ip"):
                SubElement(intf_el, "ip-address").text = intf["ip"]
            if intf.get("speed"):
                SubElement(intf_el, "speed").text = intf["speed"]
            if intf.get("mode"):
                SubElement(intf_el, "switchport-mode").text = intf["mode"]
            if intf.get("vlan"):
                SubElement(intf_el, "access-vlan").text = str(intf["vlan"])
            if intf.get("channel_group"):
                SubElement(intf_el, "channel-group").text = str(intf["channel_group"])
            if intf.get("vrf"):
                SubElement(intf_el, "vrf-member").text = intf["vrf"]
            if intf.get("mtu"):
                SubElement(intf_el, "mtu").text = str(intf["mtu"])
            shutdown_el = SubElement(intf_el, "shutdown")
            shutdown_el.text = "true" if intf.get("shutdown") else "false"

    # BGP
    if device.asn:
        bgp_el = SubElement(root, "router-bgp", asn=str(device.asn))
        bgp_cfg = device.config.get("bgp", {})
        if bgp_cfg.get("router_id"):
            SubElement(bgp_el, "router-id").text = bgp_cfg["router_id"]
        elif device.loopback0:
            SubElement(bgp_el, "router-id").text = device.loopback0.split("/")[0]

        for nbr_ip, nbr_cfg in bgp_cfg.get("neighbors", {}).items():
            nbr_el = SubElement(bgp_el, "neighbor", address=nbr_ip)
            if nbr_cfg.get("remote_as"):
                SubElement(nbr_el, "remote-as").text = str(nbr_cfg["remote_as"])
            if nbr_cfg.get("update_source"):
                SubElement(nbr_el, "update-source").text = nbr_cfg["update_source"]
            if nbr_cfg.get("ebgp_multihop"):
                SubElement(nbr_el, "ebgp-multihop").text = str(nbr_cfg["ebgp_multihop"])
            if nbr_cfg.get("peer_type"):
                SubElement(nbr_el, "peer-type").text = nbr_cfg["peer_type"]
            for af_name, af_cfg in nbr_cfg.get("address_families", {}).items():
                af_el = SubElement(nbr_el, "address-family", name=af_name)
                if af_cfg.get("send_community"):
                    SubElement(af_el, "send-community").text = af_cfg["send_community"]
                if af_cfg.get("rewrite_evpn_rt_asn"):
                    SubElement(af_el, "rewrite-evpn-rt-asn")
                if af_cfg.get("route_reflector_client"):
                    SubElement(af_el, "route-reflector-client")

    # VRFs
    vrfs = device.config.get("vrfs", {})
    if vrfs:
        vrfs_el = SubElement(root, "vrfs")
        for vrf_name, vrf_cfg in vrfs.items():
            vrf_el = SubElement(vrfs_el, "vrf", name=vrf_name)
            if vrf_cfg.get("vni"):
                SubElement(vrf_el, "vni").text = str(vrf_cfg["vni"])
            if vrf_cfg.get("rd"):
                SubElement(vrf_el, "rd").text = vrf_cfg["rd"]

    # NVE
    nve = device.config.get("nve", {})
    if nve:
        nve_el = SubElement(root, "interface-nve1")
        if nve.get("source_interface"):
            SubElement(nve_el, "source-interface").text = nve["source_interface"]
        if nve.get("host_reachability"):
            SubElement(nve_el, "host-reachability-protocol").text = nve["host_reachability"]
        if nve.get("multisite_bgw_intf"):
            SubElement(nve_el, "multisite-border-gateway-interface").text = nve["multisite_bgw_intf"]
        for vni_id, vni_cfg in nve.get("members", {}).items():
            member = SubElement(nve_el, "member-vni", id=str(vni_id))
            if vni_cfg.get("associate_vrf"):
                member.set("associate-vrf", "true")
            if vni_cfg.get("multisite_ir"):
                SubElement(member, "multisite-ingress-replication")
            if vni_cfg.get("ingress_replication"):
                SubElement(member, "ingress-replication-protocol").text = vni_cfg["ingress_replication"]

    # vPC
    if device.vpc_domain:
        vpc_el = SubElement(root, "vpc-domain", id=str(device.vpc_domain))
        vpc_cfg = device.config.get("vpc", {})
        if device.vpc_peer:
            SubElement(vpc_el, "peer-hostname").text = device.vpc_peer
        if vpc_cfg.get("peer_keepalive"):
            SubElement(vpc_el, "peer-keepalive").text = vpc_cfg["peer_keepalive"]
        if vpc_cfg.get("peer_link"):
            SubElement(vpc_el, "peer-link").text = vpc_cfg["peer_link"]
        if vpc_cfg.get("peer_gateway"):
            SubElement(vpc_el, "peer-gateway")

    raw_xml = tostring(root, encoding="unicode", xml_declaration=False)
    pretty = parseString(f'<?xml version="1.0" encoding="UTF-8"?>\n{raw_xml}').toprettyxml(indent="  ")
    lines = [l for l in pretty.split("\n") if l.strip()]
    return "\n".join(lines)


# =============================================================================
# FABRIC BUILDER - ENDPOINTS API
# =============================================================================

@app.post("/api/fabric/endpoints")
async def add_endpoint(request: Request):
    """Add an endpoint to the fabric."""
    body = await request.json()
    ep = _endpoint_store.add(body)
    return {"endpoint": ep.to_dict()}


@app.get("/api/fabric/endpoints")
async def list_endpoints():
    """List all fabric endpoints."""
    return {"endpoints": _endpoint_store.list_all()}


@app.put("/api/fabric/endpoints/{ep_id}")
async def update_endpoint(ep_id: str, request: Request):
    """Update an endpoint."""
    body = await request.json()
    ep = _endpoint_store.update(ep_id, body)
    if not ep:
        raise HTTPException(status_code=404, detail="Endpoint not found")
    return {"endpoint": ep.to_dict()}


@app.delete("/api/fabric/endpoints/{ep_id}")
async def delete_endpoint(ep_id: str):
    """Remove an endpoint."""
    removed = _endpoint_store.remove(ep_id)
    if not removed:
        raise HTTPException(status_code=404, detail="Endpoint not found")
    return {"status": "removed"}


@app.post("/api/fabric/devices")
async def add_device(request: Request):
    """Add a new switch device to the fabric."""
    import uuid as _uuid
    if not _fabric_model:
        raise HTTPException(status_code=404, detail="No fabric model loaded")
    body = await request.json()
    body.setdefault("id", str(_uuid.uuid4()))
    from fabric_builder.fabric_model import FabricDevice
    device = FabricDevice(body)
    _fabric_model.devices.append(device)
    return {"device": device.to_dict()}


@app.post("/api/fabric/links")
async def add_link(request: Request):
    """Add a new link to the fabric."""
    import uuid as _uuid
    if not _fabric_model:
        raise HTTPException(status_code=404, detail="No fabric model loaded")
    body = await request.json()
    body.setdefault("id", str(_uuid.uuid4()))

    from_id = body.get("from_device", "")
    to_id = body.get("to_device", "")
    from_dev = _fabric_model.get_device(from_id)
    to_dev = _fabric_model.get_device(to_id)
    if from_dev:
        body["from_device"] = from_dev.hostname
    if to_dev:
        body["to_device"] = to_dev.hostname

    from fabric_builder.fabric_model import FabricLink
    link = FabricLink(body)
    _fabric_model.links.append(link)
    return {"link": link.to_dict()}


# =============================================================================
# FABRIC BUILDER - TRAFFIC SIMULATION API
# =============================================================================

@app.post("/api/fabric/traffic/trace")
async def trace_traffic(request: Request):
    """Trace traffic path between two endpoints."""
    if not _fabric_model:
        raise HTTPException(status_code=404, detail="No fabric model loaded")
    _init_traffic_engine()

    body = await request.json()
    src_id = body.get("src_endpoint_id", "")
    dst_id = body.get("dst_endpoint_id", "")
    vlan = body.get("vlan", "")
    vrf = body.get("vrf", "")

    result = _traffic_engine.trace(src_id, dst_id, vlan=vlan, vrf=vrf)
    return result


@app.post("/api/fabric/traffic/failover")
async def simulate_failover(request: Request):
    """Simulate a failure and compute failover path."""
    if not _fabric_model:
        raise HTTPException(status_code=404, detail="No fabric model loaded")
    _init_traffic_engine()

    body = await request.json()
    failure = body.get("failure", {})
    f_type = failure.get("type", "link")
    f_target = failure.get("target_id", "")

    if f_type == "device":
        result = _failover_sim.simulate_device_failure(f_target)
    elif f_type == "link":
        result = _failover_sim.simulate_link_failure(f_target)
    else:
        result = {"error": "Unknown failure type", "converged": False}

    return result


@app.post("/api/fabric/traffic/restore")
async def restore_failure(request: Request):
    """Restore a previously injected failure."""
    if not _traffic_engine:
        raise HTTPException(status_code=404, detail="No traffic engine initialized")

    body = await request.json()
    f_type = body.get("type", "link")
    f_target = body.get("target_id", "")
    _traffic_engine.restore(f_type, f_target)
    return {"status": "restored"}


def _detect_hostname(filename: str, text: str) -> str:
    """Detect hostname from file content or filename."""
    hostname = extract_hostname(text)
    if hostname:
        return hostname

    stem = Path(filename).stem
    noise_words = (
        r"show|config|output|commands?|backup|running|startup|"
        r"cdp|lldp|neighbors?|detail|brief|interfaces?|status|"
        r"version|inventory|platform|ip|route|bgp|ospf|vlan|"
        r"description|etherchannel|port-?channel|spanning-?tree|"
        r"log|tech-?support|diag"
    )
    stem_clean = re.sub(
        r"[-_.]?(" + noise_words + r")[-_.]?", " ", stem, flags=re.IGNORECASE
    )
    stem_clean = stem_clean.strip(" -_.")
    stem_clean = re.sub(r"\s+", " ", stem_clean).strip()
    if stem_clean and len(stem_clean) >= 2 and not stem_clean.isdigit():
        return stem_clean

    return ""


def _split_multi_device_output(text: str) -> list[tuple[str, str]]:
    """
    Detect if a file contains outputs from multiple devices
    (e.g. collected via jump host or automation tool).
    """
    device_markers = re.findall(
        r"^[!=]{3,}\s*(\S+)\s*[!=]{3,}$|^#{3,}\s*(\S+)\s*#{3,}$|^---\s*(\S+)\s*---$",
        text, re.MULTILINE
    )

    if not device_markers:
        return []

    sections = []
    pattern = r"(?:^[!=]{3,}\s*(\S+)\s*[!=]{3,}$|^#{3,}\s*(\S+)\s*#{3,}$|^---\s*(\S+)\s*---$)"
    splits = re.split(pattern, text, flags=re.MULTILINE)

    current_host = ""
    for part in splits:
        if part is None:
            continue
        part = part.strip()
        if not part:
            continue
        if len(part) < 50 and re.match(r"^[\w.-]+$", part):
            current_host = part
        elif current_host:
            sections.append((current_host, part))

    return sections


def _parse_device_output(builder: TopologyBuilder, hostname: str, text: str,
                         routing_builder: RoutingTopologyBuilder = None,
                         known_hostnames: set = None) -> list[str]:
    """Parse all recognizable command outputs for a device. Returns list of commands found."""
    builder.add_device(hostname)

    sections = _identify_command_sections(text)
    commands_found = []

    all_neighbors = []
    has_cdp_lldp = False

    for section_type, section_text in sections:
        if section_type == "cdp":
            neighbors = parse_cdp_neighbors(section_text, hostname)
            all_neighbors.extend(neighbors)
            has_cdp_lldp = True
            commands_found.append("CDP neighbors")

        elif section_type == "lldp":
            neighbors = parse_lldp_neighbors(section_text, hostname)
            all_neighbors.extend(neighbors)
            has_cdp_lldp = True
            commands_found.append("LLDP neighbors")

        elif section_type == "interface_brief":
            interfaces = parse_interface_brief(section_text, hostname)
            builder.add_interfaces(hostname, interfaces)
            commands_found.append("interface brief")

        elif section_type == "interface_description":
            descriptions = parse_interface_description(section_text, hostname)
            if not has_cdp_lldp:
                inferred = infer_neighbors_from_descriptions(descriptions)
                inferred = _filter_inferred_by_known(inferred, known_hostnames)
                all_neighbors.extend(inferred)
            commands_found.append("interface description")

        elif section_type == "interface_status":
            statuses = parse_interface_status(section_text, hostname)
            if not has_cdp_lldp:
                inferred = infer_neighbors_from_descriptions(statuses)
                inferred = _filter_inferred_by_known(inferred, known_hostnames)
                all_neighbors.extend(inferred)
            commands_found.append("interface status")

        elif section_type == "platform":
            device_info = parse_platform_detail(section_text, hostname)
            builder.add_device(hostname, device_info)
            commands_found.append("platform/version")

        elif section_type == "running_config":
            config = parse_running_config(section_text)
            if config.get("hostname"):
                builder.add_device(config["hostname"])
            device_info = {
                "hostname": config.get("hostname", hostname),
                "config": config,
            }
            builder.add_device(hostname, device_info)
            commands_found.append("running-config")

            if routing_builder:
                bgp_cfg = parse_bgp_from_config(section_text)
                if bgp_cfg.get("local_asn"):
                    routing_builder.add_bgp_config(hostname, bgp_cfg)
                    commands_found.append("BGP (from config)")
                ospf_cfg = parse_ospf_from_config(section_text)
                if ospf_cfg.get("process_id"):
                    routing_builder.add_ospf_config(hostname, ospf_cfg)
                    commands_found.append("OSPF (from config)")

        elif section_type == "bgp_summary":
            if routing_builder:
                summary = parse_bgp_summary(section_text, hostname)
                routing_builder.add_bgp_summary(hostname, summary)
                commands_found.append("BGP summary")

        elif section_type == "bgp_neighbors":
            if routing_builder:
                peers = parse_bgp_neighbors_detail(section_text, hostname)
                routing_builder.add_bgp_neighbor_detail(hostname, peers)
                commands_found.append("BGP neighbors detail")

        elif section_type == "ospf_overview":
            if routing_builder:
                overview = parse_ospf_overview(section_text, hostname)
                routing_builder.add_ospf_overview(hostname, overview)
                commands_found.append("OSPF overview")

        elif section_type == "ospf_neighbors":
            if routing_builder:
                nbrs = parse_ospf_neighbors(section_text, hostname)
                routing_builder.add_ospf_neighbors(hostname, nbrs)
                commands_found.append("OSPF neighbors")

        elif section_type == "ospf_interfaces":
            if routing_builder:
                intfs = parse_ospf_interfaces(section_text, hostname)
                routing_builder.add_ospf_interfaces(hostname, intfs)
                commands_found.append("OSPF interfaces")

    if not sections:
        neighbors = parse_cdp_neighbors(text, hostname)
        if neighbors:
            all_neighbors.extend(neighbors)
            has_cdp_lldp = True
            commands_found.append("CDP neighbors")

        neighbors = parse_lldp_neighbors(text, hostname)
        if neighbors:
            all_neighbors.extend(neighbors)
            has_cdp_lldp = True
            commands_found.append("LLDP neighbors")

        if not has_cdp_lldp:
            descriptions = parse_interface_description(text, hostname)
            if descriptions:
                inferred = infer_neighbors_from_descriptions(descriptions)
                inferred = _filter_inferred_by_known(inferred, known_hostnames)
                all_neighbors.extend(inferred)
                commands_found.append("interface description (inferred)")

        device_info = parse_platform_detail(text, hostname)
        if device_info.get("model"):
            builder.add_device(hostname, device_info)
            commands_found.append("platform/version")

    if all_neighbors:
        builder.add_neighbors(all_neighbors)

    return commands_found


def _filter_inferred_by_known(inferred: list[dict], known_hostnames: set = None) -> list[dict]:
    """
    Filter inferred neighbors to only include those whose remote_device
    matches a known hostname from the uploaded file bundle.
    """
    if not known_hostnames or not inferred:
        return inferred

    known_lower = {h.lower() for h in known_hostnames}
    filtered = []
    for n in inferred:
        remote = n.get("remote_device", "").lower()
        if remote in known_lower:
            filtered.append(n)
        else:
            for kh in known_lower:
                if remote in kh or kh in remote:
                    filtered.append(n)
                    break
    return filtered


def _identify_command_sections(text: str) -> list[tuple[str, str]]:
    """
    Identify different command output sections within a file.
    Looks for 'show' command prompts or markers.
    """
    sections = []

    show_pattern = re.compile(
        r"^(\S+)[#>]\s*(show\s+.+)$",
        re.MULTILINE | re.IGNORECASE
    )

    matches = list(show_pattern.finditer(text))

    if not matches:
        section_type = _classify_text_content(text)
        if section_type:
            return [(section_type, text)]
        return []

    for i, match in enumerate(matches):
        command = match.group(2).lower().strip()
        start = match.start()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        section_text = text[start:end]

        section_type = _classify_command(command)
        if section_type:
            sections.append((section_type, section_text))

    return sections


def _classify_command(command: str) -> str:
    """Classify a show command into a parser category."""
    cmd = command.lower()

    if "cdp" in cmd and "neighbor" in cmd:
        return "cdp"
    if "lldp" in cmd and "neighbor" in cmd:
        return "lldp"
    if "interface" in cmd and "brief" in cmd:
        return "interface_brief"
    if "interface" in cmd and "description" in cmd:
        return "interface_description"
    if "interface" in cmd and "status" in cmd:
        return "interface_status"

    if "bgp" in cmd and "summary" in cmd:
        return "bgp_summary"
    if "bgp" in cmd and "neighbor" in cmd:
        return "bgp_neighbors"
    if "ospf" in cmd and "neighbor" in cmd:
        return "ospf_neighbors"
    if "ospf" in cmd and "interface" in cmd:
        return "ospf_interfaces"
    if "ospf" in cmd and ("overview" in cmd or cmd.strip().endswith("ospf")):
        return "ospf_overview"

    if any(k in cmd for k in ["platform", "inventory", "version", "module"]):
        return "platform"
    if any(k in cmd for k in ["running", "run", "startup"]):
        return "running_config"
    if "etherchannel" in cmd or "port-channel" in cmd or "lacp" in cmd:
        return "running_config"
    if "vpc" in cmd:
        return "running_config"
    if "cdp" in cmd:
        return "cdp"
    if "lldp" in cmd:
        return "lldp"
    if "bgp" in cmd:
        return "bgp_summary"
    if "ospf" in cmd:
        return "ospf_overview"

    return ""


def _classify_text_content(text: str) -> str:
    """Classify text content by looking for characteristic patterns across vendors."""
    text_lower = text[:3000].lower()

    # CDP detection
    if "device id" in text_lower and ("local intrfce" in text_lower or "interface" in text_lower):
        if "cdp" in text_lower or "holdtme" in text_lower or "port id" in text_lower:
            return "cdp"

    # LLDP detection (Cisco, Arista, Juniper)
    if any(k in text_lower for k in ["chassis id", "system name", "lldp neighbor"]):
        if any(k in text_lower for k in ["lldp", "port id", "port description", "system description"]):
            return "lldp"

    # BGP summary detection
    if re.search(r"bgp\s+router\s+identifier", text_lower):
        if re.search(r"neighbor\s+v\s+as", text_lower):
            return "bgp_summary"
    if "bgp state" in text_lower and "remote as" in text_lower:
        return "bgp_neighbors"

    # OSPF detection
    if re.search(r"ospf\s+router\s+with\s+id|ospf\s+process|routing process", text_lower):
        if re.search(r"neighbor\s+id", text_lower):
            return "ospf_neighbors"
        return "ospf_overview"
    if re.search(r"neighbor\s+id\s+.*?(?:state|pri)", text_lower) and "ospf" in text_lower:
        return "ospf_neighbors"

    # Interface brief (IOS/NX-OS/Arista)
    if re.search(r"interface\s+ip.address\s+ok\?", text_lower):
        return "interface_brief"
    if re.search(r"interface\s+ip\s+address\s+status", text_lower):
        return "interface_brief"

    # Interface description
    if "description" in text_lower and re.search(r"interface\s+status\s+protocol", text_lower):
        return "interface_description"

    # Interface status (NX-OS / IOS)
    if re.search(r"port\s+name\s+status\s+vlan", text_lower):
        return "interface_status"

    # Running config detection (multi-vendor)
    if "hostname" in text_lower and "interface" in text_lower and "!" in text:
        return "running_config"
    if "switchname" in text_lower and "interface" in text_lower:
        return "running_config"
    # Juniper set-style config
    if "set system host-name" in text_lower or "set interfaces" in text_lower:
        return "running_config"

    # Platform/version
    if any(k in text_lower for k in [
        "pid:", "serial number", "model number", "uptime is",
        "system version", "eos version", "junos", "pan-os",
        "fortigate", "big-ip", "software image version"
    ]):
        return "platform"

    return ""


# =============================================================================
# NEXUS DASHBOARD INTEGRATION
# =============================================================================

@app.post("/api/nd/authenticate")
async def nd_authenticate(request: Request):
    """Authenticate to Nexus Dashboard and return a session token."""
    body = await request.json()
    url = body.get("url", "").rstrip("/")
    username = body.get("username", "")
    password = body.get("password", "")

    if not url or not username or not password:
        raise HTTPException(status_code=400, detail="URL, username, and password are required")

    login_url = f"{url}/login"
    payload = {"userName": username, "userPasswd": password, "domain": "local"}

    try:
        async with httpx.AsyncClient(verify=False, timeout=15.0) as client:
            resp = await client.post(login_url, json=payload)
            if resp.status_code == 200:
                data = resp.json()
                token = data.get("token") or data.get("jwttoken") or ""
                if not token and "Dcnm-Token" in resp.headers:
                    token = resp.headers["Dcnm-Token"]
                if not token:
                    token = resp.cookies.get("AuthCookie", "")
                if not token:
                    raise HTTPException(status_code=401, detail="Authentication succeeded but no token received")
                return {"token": token, "message": "Authenticated successfully"}
            else:
                detail = "Authentication failed"
                try:
                    err = resp.json()
                    detail = err.get("message", err.get("error", detail))
                except Exception:
                    pass
                raise HTTPException(status_code=resp.status_code, detail=detail)
    except httpx.ConnectError:
        raise HTTPException(status_code=502, detail=f"Cannot connect to {url}")
    except httpx.TimeoutException:
        raise HTTPException(status_code=504, detail="Connection timed out")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/nd/push-config")
async def nd_push_config(request: Request):
    """Push device configuration(s) to Nexus Dashboard Fabric Controller."""
    if not _fabric_model:
        raise HTTPException(status_code=404, detail="No fabric model loaded")

    body = await request.json()
    token = body.get("token", "")
    url = body.get("url", "").rstrip("/")
    scope = body.get("scope", "all")

    if not token or not url:
        raise HTTPException(status_code=400, detail="Token and URL required")

    headers = {
        "Dcnm-Token": token,
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json"
    }

    if scope == "all":
        devices = _fabric_model.devices
    else:
        device = _fabric_model.get_device(scope)
        if not device:
            raise HTTPException(status_code=404, detail=f"Device {scope} not found")
        devices = [device]

    results = []
    config_engine = None
    if _fabric_model:
        try:
            from fabric_builder.config_engine import ConfigEngine
            config_engine = ConfigEngine(_fabric_model)
        except Exception:
            pass

    async with httpx.AsyncClient(verify=False, timeout=30.0) as client:
        for device in devices:
            cli_config = ""
            if config_engine:
                try:
                    cli_config = config_engine.get_device_config(device.hostname)
                except Exception:
                    cli_config = ""

            deploy_payload = {
                "serialNumber": device.hostname,
                "hostname": device.hostname,
                "managementIpAddress": device.mgmt_ip.split("/")[0] if device.mgmt_ip else "",
                "model": device.model or "N9K-C93180YC-FX3",
                "role": device.role.replace("_", " ").title(),
                "config": cli_config
            }

            try:
                deploy_url = f"{url}/appcenter/cisco/ndfc/api/v1/lan-fabric/rest/control/fabrics/default/config-deploy"
                resp = await client.post(deploy_url, json=deploy_payload, headers=headers)
                if resp.status_code in (200, 201, 202):
                    results.append({"device": device.hostname, "status": "success"})
                else:
                    detail = resp.text[:200] if resp.text else "Unknown error"
                    results.append({"device": device.hostname, "status": "failed", "detail": detail})
            except Exception as e:
                results.append({"device": device.hostname, "status": "error", "detail": str(e)})

    success_count = sum(1 for r in results if r["status"] == "success")
    total = len(results)
    return {
        "message": f"Pushed {success_count}/{total} device(s) successfully",
        "results": results
    }


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)

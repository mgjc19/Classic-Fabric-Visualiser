"""
BOM Parser - Auto-detects columns from Excel/CSV and normalizes into fabric device/link data.
Supports two modes:
  1. Hardware BOM (PIDs, quantities, SFPs, cables) → extracted into hardware inventory
  2. Fabric BOM (hostnames, roles, IPs) → directly into device model
"""
import io
import re
import uuid
from pathlib import Path
from typing import Any

import openpyxl


ROLE_ALIASES = {
    "spine": "spine",
    "sp": "spine",
    "super_spine": "super_spine",
    "super spine": "super_spine",
    "superspine": "super_spine",
    "ssp": "super_spine",
    "leaf": "leaf",
    "lf": "leaf",
    "compute_leaf": "leaf",
    "compute leaf": "leaf",
    "border_leaf": "border_leaf",
    "border leaf": "border_leaf",
    "bleaf": "border_leaf",
    "bl": "border_leaf",
    "border_gateway": "border_gateway",
    "border gateway": "border_gateway",
    "bgw": "border_gateway",
    "service_leaf": "service_leaf",
    "service leaf": "service_leaf",
    "svc_leaf": "service_leaf",
    "oob_switch": "oob_switch",
    "oob switch": "oob_switch",
    "oob": "oob_switch",
    "mgmt": "oob_switch",
}

NEXUS_MODELS = {
    # Super-spine / high-radix spine candidates (400G, 64+ ports)
    "N9K-C9364C-GX": {"role": "spine", "tier": "super_spine", "ports_100g": 64, "ports_400g": 64, "description": "Nexus 9364C-GX 64p 100/400G"},
    "N9K-C9332D-GX2B": {"role": "spine", "tier": "super_spine", "ports_100g": 32, "ports_400g": 32, "description": "Nexus 9332D-GX2B 32p 400G"},
    "N9K-C9364D-GX2A": {"role": "spine", "tier": "super_spine", "ports_100g": 64, "ports_400g": 64, "description": "Nexus 9364D-GX2A 64p 400G"},
    "N9K-C9408": {"role": "spine", "tier": "super_spine", "ports_100g": 0, "ports_400g": 0, "modular": True, "description": "Nexus 9408 Modular Chassis (Super-Spine)"},
    "N9K-C9508": {"role": "spine", "tier": "super_spine", "ports_100g": 0, "ports_400g": 0, "modular": True, "description": "Nexus 9508 Modular Chassis (Super-Spine)"},
    "N9K-C9516": {"role": "spine", "tier": "super_spine", "ports_100g": 0, "ports_400g": 0, "modular": True, "description": "Nexus 9516 Modular Chassis (Super-Spine)"},
    # Standard spine (36p or 32p 100G)
    "N9K-C9336C-FX2": {"role": "spine", "tier": "spine", "ports_100g": 36, "description": "Nexus 9336C-FX2 36p 100G"},
    "N9K-C9332C": {"role": "spine", "tier": "spine", "ports_100g": 32, "description": "Nexus 9332C 32p 100G"},
    "N9K-C9364C": {"role": "spine", "tier": "spine", "ports_100g": 64, "description": "Nexus 9364C 64p 100G"},
    # Leaf switches
    "N9K-C93180YC-FX": {"role": "leaf", "tier": "leaf", "ports_10g": 48, "ports_100g": 6, "description": "Nexus 93180YC-FX 48p 10/25G + 6p 100G"},
    "N9K-C93180YC-FX3": {"role": "leaf", "tier": "leaf", "ports_10g": 48, "ports_100g": 6, "description": "Nexus 93180YC-FX3 48p 10/25G + 6p 100G"},
    "N9K-C93180YC-FX3S": {"role": "leaf", "tier": "leaf", "ports_10g": 48, "ports_100g": 6, "description": "Nexus 93180YC-FX3S 48p 10/25G + 6p 100G"},
    "N9K-C93180YC-EX": {"role": "leaf", "tier": "leaf", "ports_10g": 48, "ports_100g": 6, "description": "Nexus 93180YC-EX 48p 10/25G + 6p 100G"},
    "N9K-C93108TC-FX": {"role": "leaf", "tier": "leaf", "ports_1g": 48, "ports_100g": 6, "description": "Nexus 93108TC-FX 48p 1G + 6p 100G"},
    "N9K-C93108TC-FX3P": {"role": "leaf", "tier": "leaf", "ports_1g": 48, "ports_100g": 6, "description": "Nexus 93108TC-FX3P 48p 1G + 6p 100G"},
    "N9K-C93240YC-FX2": {"role": "leaf", "tier": "leaf", "ports_10g": 48, "ports_100g": 12, "description": "Nexus 93240YC-FX2 48p 25G + 12p 100G"},
    "N9K-C93360YC-FX2": {"role": "leaf", "tier": "leaf", "ports_10g": 96, "ports_100g": 12, "description": "Nexus 93360YC-FX2 96p 25G + 12p 100G"},
    "N9K-C9348GC-FXP": {"role": "leaf", "tier": "leaf", "ports_1g": 48, "ports_10g": 4, "description": "Nexus 9348GC-FXP 48p 1G + 4p 10G"},
    "N9K-C9372PX": {"role": "leaf", "tier": "leaf", "ports_10g": 48, "ports_100g": 6, "description": "Nexus 9372PX 48p 10G + 6p 40G"},
}

SFP_TYPES = {
    "QSFP-100G-SR4-S": {"speed": "100G", "type": "MMF", "reach": "100m"},
    "QSFP-100G-PSM4-S": {"speed": "100G", "type": "SMF", "reach": "500m"},
    "QSFP-100G-LR4-S": {"speed": "100G", "type": "SMF", "reach": "10km"},
    "QSFP-100G-CU1M": {"speed": "100G", "type": "DAC", "reach": "1m"},
    "QSFP-100G-CU2M": {"speed": "100G", "type": "DAC", "reach": "2m"},
    "QSFP-100G-CU3M": {"speed": "100G", "type": "DAC", "reach": "3m"},
    "QSFP-100G-AOC": {"speed": "100G", "type": "AOC", "reach": "variable"},
    "QSFP-40G-SR4": {"speed": "40G", "type": "MMF", "reach": "100m"},
    "QSFP-40G-LR4": {"speed": "40G", "type": "SMF", "reach": "10km"},
    "SFP-25G-SR-S": {"speed": "25G", "type": "MMF", "reach": "100m"},
    "SFP-10G-SR-S": {"speed": "10G", "type": "MMF", "reach": "300m"},
    "SFP-10G-LR-S": {"speed": "10G", "type": "SMF", "reach": "10km"},
}

COLUMN_PATTERNS_HARDWARE = {
    "pid": r"product[\s_.-]*id|part[\s_.-]*num|^pid$|^sku$|^model$|cisco[\s_.-]*p|item[\s_.-]*(id|code|no|num)",
    "description": r"desc|product[\s_.-]*desc|item[\s_.-]*desc",
    "quantity": r"^qty$|^quantity$|^count$|^num$|^units?$|^ordered|order[\s_.-]*qty",
    "unit_price": r"price|cost|unit[\s_.-]*price|list[\s_.-]*price",
    "role": r"^role$|^function$|^purpose$|^use$",
}

COLUMN_PATTERNS_FABRIC = {
    "hostname": r"^host\s*name$|^device[\s_.-]*name$|^switch[\s_.-]*name$|^equipment[\s_.-]*name$|^node[\s_.-]*name$|^hostname$|^device$|^switch$|^host$|^node$|^equipment$|^name$",
    "role": r"^role$|^function$|^device[\s_.-]*type$|^type$|^tier$|^layer$",
    "model": r"model|platform|part[\s_.-]*num|pid|sku|hardware",
    "serial": r"serial|^sn$|serial[\s_.-]*num",
    "mgmt_ip": r"mgmt[\s_.-]*ip|management[\s_.-]*ip|oob[\s_.-]*ip|ip[\s_.-]*addr|^ip$",
    "loopback0": r"loopback[\s_.-]*0|^lo0$|router[\s_.-]*id|^rid$",
    "loopback1": r"loopback[\s_.-]*1|^lo1$|vtep|nve[\s_.-]*ip",
    "site": r"^site$|^location$|^dc$|^datacenter$|data[\s_.-]*center|^pod$|^fabric$",
    "vpc_domain": r"vpc[\s_.-]*domain|vpc[\s_.-]*id|mct[\s_.-]*id",
    "vpc_peer": r"vpc[\s_.-]*peer|mct[\s_.-]*peer|peer[\s_.-]*switch",
    "sfp": r"sfp|transceiver|optic|xcvr|module",
    "cable_type": r"cable[\s_.-]*type|^media$|^fiber$|^cable$|media[\s_.-]*type",
    "connected_to": r"connect[\w]*[\s_.-]*to|remote[\s_.-]*device|^neighbor$|^peer$|^remote$|^dest|^to[\s_.-]*device|^far[\s_.-]*end",
    "local_port": r"local[\s_.-]*port|local[\s_.-]*int|^port$|^interface$|from[\s_.-]*port|^src[\s_.-]*port|^a[\s_.-]*port",
    "remote_port": r"remote[\s_.-]*port|remote[\s_.-]*int|to[\s_.-]*port|peer[\s_.-]*port|^dest[\s_.-]*port|^b[\s_.-]*port|^far[\s_.-]*port",
    "speed": r"speed|bandwidth|^bw$|link[\s_.-]*speed|^rate$",
    "quantity": r"^qty$|^quantity$|^count$|^num$",
}


class BomParser:
    """Parses BOM Excel/CSV files with auto-column-detection.
    
    Supports two BOM types:
    - Hardware BOM: PIDs, quantities, SFPs, cables (no hostnames)
    - Fabric BOM: Hostnames, roles, IPs, ports (ready-to-use)
    """

    def parse(self, content: bytes, filename: str = "") -> dict[str, Any]:
        """
        Parse a BOM file and return structured data.
        Auto-detects whether it's a hardware BOM or a fabric BOM.
        Returns: {"type": "hardware"|"fabric", "devices": [...], "links": [...], "hardware": {...}, "metadata": {...}}
        """
        filename_lower = filename.lower()
        if filename_lower.endswith((".xlsx", ".xls")):
            return self._parse_excel(content)
        elif filename_lower.endswith(".csv"):
            return self._parse_csv(content)
        else:
            return self._parse_excel(content)

    def _parse_excel(self, content: bytes) -> dict[str, Any]:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)

        all_rows = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                all_rows[sheet_name] = rows

        wb.close()

        if not all_rows:
            return self._empty_result("hardware")

        bom_type = self._detect_bom_type(all_rows)

        if bom_type == "hardware":
            return self._parse_hardware_bom(all_rows)
        else:
            return self._parse_fabric_bom(all_rows)

    def _parse_csv(self, content: bytes) -> dict[str, Any]:
        import csv
        text = content.decode("utf-8", errors="replace")
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
        if not rows:
            return self._empty_result("hardware")

        all_rows = {"Sheet1": [tuple(r) for r in rows]}
        bom_type = self._detect_bom_type(all_rows)

        if bom_type == "hardware":
            return self._parse_hardware_bom(all_rows)
        else:
            return self._parse_fabric_bom(all_rows)

    def _detect_bom_type(self, all_rows: dict[str, list[tuple]]) -> str:
        """Detect whether this is a hardware BOM (PIDs/quantities) or a fabric BOM (hostnames/IPs)."""
        for sheet_name, rows in all_rows.items():
            header_idx, headers = self._find_header_row(rows)
            headers_lower = " ".join(h.lower() for h in headers if h)

            if any(k in headers_lower for k in ["hostname", "switch name", "device name", "loopback", "mgmt ip", "management ip"]):
                return "fabric"

            has_pid_indicators = False
            for row in rows[header_idx + 1: header_idx + 20]:
                row_text = " ".join(str(c) for c in row if c).upper()
                if re.search(r"N9K-|QSFP-|SFP-|GLC-|FET-|CAB-|FS-", row_text):
                    has_pid_indicators = True
                    break

            if has_pid_indicators:
                return "hardware"

            if any(k in headers_lower for k in ["qty", "quantity", "price", "cost", "pid", "part num", "product id"]):
                return "hardware"

        return "hardware"

    def _parse_hardware_bom(self, all_rows: dict[str, list[tuple]]) -> dict[str, Any]:
        """Parse a hardware BOM with PIDs and quantities into a hardware inventory."""
        switches = []
        sfps = []
        cables = []
        other_items = []

        for sheet_name, rows in all_rows.items():
            header_idx, headers = self._find_header_row(rows)
            col_map = self._detect_columns_hardware(headers)

            for row in rows[header_idx + 1:]:
                if not any(cell for cell in row if cell):
                    continue

                record = {}
                for col_idx, cell in enumerate(row):
                    if col_idx < len(row) and col_idx in col_map:
                        record[col_map[col_idx]] = str(cell).strip() if cell else ""

                if not record.get("pid"):
                    for col_idx, cell in enumerate(row):
                        if cell:
                            cell_str = str(cell).strip().upper()
                            if re.match(r"N9K-|QSFP-|SFP-|GLC-|FET-|CAB-|FS-", cell_str):
                                record["pid"] = cell_str
                                break

                if not record.get("pid"):
                    continue

                pid = record["pid"].upper().strip()
                qty = self._parse_qty(record.get("quantity", "1"))
                desc = record.get("description", "")
                role_hint = record.get("role", "")

                item = {"pid": pid, "quantity": qty, "description": desc, "role_hint": role_hint}

                if any(pid.startswith(p) for p in ["N9K-", "N3K-", "N5K-", "N7K-"]):
                    model_info = NEXUS_MODELS.get(pid, {})
                    item["model_info"] = model_info
                    item["inferred_role"] = model_info.get("role", self._infer_role_from_pid(pid, role_hint))
                    switches.append(item)
                elif any(pid.startswith(p) for p in ["QSFP-", "SFP-", "GLC-", "FET-"]):
                    sfp_info = SFP_TYPES.get(pid, {})
                    item["sfp_info"] = sfp_info
                    sfps.append(item)
                elif any(k in pid.lower() for k in ["cab-", "fs-", "fiber", "cable", "dac", "aoc", "cord"]):
                    cables.append(item)
                else:
                    other_items.append(item)

        total_switches = sum(s["quantity"] for s in switches)
        total_sfps = sum(s["quantity"] for s in sfps)
        total_cables = sum(c["quantity"] for c in cables)

        return {
            "type": "hardware",
            "devices": [],
            "links": [],
            "hardware": {
                "switches": switches,
                "sfps": sfps,
                "cables": cables,
                "other": other_items,
                "summary": {
                    "total_switches": total_switches,
                    "total_sfps": total_sfps,
                    "total_cables": total_cables,
                    "switch_models": list(set(s["pid"] for s in switches)),
                }
            },
            "metadata": {
                "total_devices": total_switches,
                "total_links": 0,
                "sites": ["site-1"],
                "multisite": False,
                "bom_type": "hardware",
            }
        }

    def _parse_fabric_bom(self, all_rows: dict[str, list[tuple]]) -> dict[str, Any]:
        """Parse a fabric BOM with hostnames and roles directly into device model."""
        devices = []
        links = []

        for sheet_name, rows in all_rows.items():
            header_idx, headers = self._find_header_row(rows)
            col_map = self._detect_columns_fabric(headers)

            has_hostname = "hostname" in col_map.values()
            has_connected_to = "connected_to" in col_map.values()

            for row in rows[header_idx + 1:]:
                if not any(cell for cell in row if cell):
                    continue
                record = {}
                for col_idx, cell in enumerate(row):
                    if col_idx in col_map:
                        record[col_map[col_idx]] = str(cell).strip() if cell else ""

                if record.get("hostname"):
                    devices.append(self._build_device(record))
                if record.get("connected_to") and record.get("hostname"):
                    links.append(self._build_link(record))

        devices = self._deduplicate_devices(devices)
        sites = set(d.get("site", "") for d in devices if d.get("site"))

        return {
            "type": "fabric",
            "devices": devices,
            "links": links,
            "hardware": None,
            "metadata": {
                "total_devices": len(devices),
                "total_links": len(links),
                "sites": sorted(sites) if sites else ["site-1"],
                "multisite": len(sites) > 1,
                "bom_type": "fabric",
            }
        }

    def _detect_columns_hardware(self, headers: list[str]) -> dict[int, str]:
        """Detect columns for hardware BOM format."""
        col_map = {}
        headers_lower = [h.lower().strip() for h in headers]

        for col_idx, header in enumerate(headers_lower):
            if not header:
                continue
            for field_name, pattern in COLUMN_PATTERNS_HARDWARE.items():
                if re.search(pattern, header, re.IGNORECASE):
                    if field_name not in col_map.values():
                        col_map[col_idx] = field_name
                        break

        return col_map

    def _detect_columns_fabric(self, headers: list[str]) -> dict[int, str]:
        """Detect columns for fabric BOM format."""
        col_map = {}
        headers_lower = [h.lower().strip() for h in headers]

        for col_idx, header in enumerate(headers_lower):
            if not header:
                continue
            for field_name, pattern in COLUMN_PATTERNS_FABRIC.items():
                if re.search(pattern, header, re.IGNORECASE):
                    if field_name not in col_map.values():
                        col_map[col_idx] = field_name
                        break

        if "hostname" not in col_map.values():
            for col_idx, header in enumerate(headers_lower):
                if not header or col_idx in col_map:
                    continue
                if any(k in header for k in ["name", "host", "device", "switch", "node"]):
                    col_map[col_idx] = "hostname"
                    break

        if "hostname" not in col_map.values() and headers_lower:
            for col_idx, header in enumerate(headers_lower):
                if header and col_idx not in col_map:
                    col_map[col_idx] = "hostname"
                    break

        return col_map

    def _find_header_row(self, rows: list[tuple]) -> tuple[int, list[str]]:
        """Find the actual header row, skipping title/blank rows."""
        for idx, row in enumerate(rows[:10]):
            cells = [str(c).strip() if c else "" for c in row]
            non_empty = [c for c in cells if c and c.lower() != "none"]
            if len(non_empty) >= 2:
                cells_lower = " ".join(non_empty).lower()
                if any(k in cells_lower for k in [
                    "name", "host", "device", "switch", "model", "role",
                    "type", "port", "interface", "serial", "ip", "cable",
                    "site", "location", "speed", "connected", "pid", "sku",
                    "equipment", "node", "spine", "leaf", "qty", "quantity",
                    "price", "description", "part", "product", "item",
                ]):
                    return idx, cells
            if len(non_empty) >= 3 and idx > 0:
                return idx, cells
        return 0, [str(c).strip() if c else "" for c in rows[0]]

    def _infer_role_from_pid(self, pid: str, role_hint: str) -> str:
        """Infer device role from PID naming patterns."""
        if role_hint:
            role_hint_lower = role_hint.lower()
            if role_hint_lower in ROLE_ALIASES:
                return ROLE_ALIASES[role_hint_lower]

        model_info = NEXUS_MODELS.get(pid.upper(), {})
        if model_info.get("tier") == "super_spine":
            return "super_spine"

        pid_upper = pid.upper()
        if "9508" in pid_upper or "9516" in pid_upper or "9408" in pid_upper:
            return "super_spine"
        if "9364" in pid_upper and ("GX" in pid_upper or "D-" in pid_upper):
            return "super_spine"
        if "9332D" in pid_upper:
            return "super_spine"
        if "9336" in pid_upper or "9332C" in pid_upper or "9364C" in pid_upper:
            return "spine"
        if "93180" in pid_upper or "93108" in pid_upper or "93240" in pid_upper or "93360" in pid_upper or "9348" in pid_upper:
            return "leaf"
        return "leaf"

    def _parse_qty(self, qty_str: str) -> int:
        """Parse quantity, handling various formats."""
        if not qty_str:
            return 1
        try:
            cleaned = re.sub(r"[^\d.]", "", qty_str)
            return max(1, int(float(cleaned))) if cleaned else 1
        except (ValueError, TypeError):
            return 1

    def _build_device(self, record: dict) -> dict:
        role_raw = record.get("role", "").lower().strip()
        role = ROLE_ALIASES.get(role_raw, self._infer_role_from_hostname(record.get("hostname", "")))

        return {
            "id": str(uuid.uuid4()),
            "hostname": record.get("hostname", ""),
            "role": role,
            "model": record.get("model", ""),
            "serial": record.get("serial", ""),
            "mgmt_ip": record.get("mgmt_ip", ""),
            "loopback0": record.get("loopback0", ""),
            "loopback1": record.get("loopback1", ""),
            "site": record.get("site", "site-1"),
            "vpc_domain": record.get("vpc_domain", ""),
            "vpc_peer": record.get("vpc_peer", ""),
            "asn": "",
            "interfaces": [],
            "config": {},
        }

    def _build_link(self, record: dict) -> dict:
        return {
            "id": str(uuid.uuid4()),
            "from_device": record.get("hostname", ""),
            "from_port": record.get("local_port", ""),
            "to_device": record.get("connected_to", ""),
            "to_port": record.get("remote_port", ""),
            "sfp": record.get("sfp", ""),
            "cable_type": record.get("cable_type", ""),
            "speed": record.get("speed", ""),
        }

    def _infer_role_from_hostname(self, hostname: str) -> str:
        hn = hostname.lower()
        if any(k in hn for k in ["spine", "sp-", "sp0"]):
            return "spine"
        if any(k in hn for k in ["bgw", "border-gw", "dci"]):
            return "border_gateway"
        if any(k in hn for k in ["bleaf", "bl-", "border"]):
            return "border_leaf"
        if any(k in hn for k in ["leaf", "lf-", "lf0", "compute"]):
            return "leaf"
        if any(k in hn for k in ["oob", "mgmt"]):
            return "oob_switch"
        return "leaf"

    def _deduplicate_devices(self, devices: list[dict]) -> list[dict]:
        seen = {}
        result = []
        for d in devices:
            hostname = d["hostname"]
            if hostname not in seen:
                seen[hostname] = d
                result.append(d)
            else:
                existing = seen[hostname]
                for key, val in d.items():
                    if val and not existing.get(key):
                        existing[key] = val
        return result

    def _empty_result(self, bom_type: str) -> dict[str, Any]:
        return {
            "type": bom_type,
            "devices": [],
            "links": [],
            "hardware": None,
            "metadata": {
                "total_devices": 0,
                "total_links": 0,
                "sites": ["site-1"],
                "multisite": False,
                "bom_type": bom_type,
            }
        }

    @staticmethod
    def generate_devices_from_hardware(hardware: dict, config: dict = None) -> dict[str, Any]:
        """
        Generate fabric devices from a hardware BOM inventory.
        Detects super-spine topology automatically based on:
        - Presence of high-radix/400G spine PIDs alongside standard spines
        - Total leaf count exceeding spine port capacity
        - Multiple distinct spine-class models with different port densities
        """
        if config is None:
            config = {}

        site = config.get("site", "DC1")
        spine_prefix = config.get("spine_prefix", f"{site}-SPINE")
        sspine_prefix = config.get("sspine_prefix", f"{site}-SSPINE")
        leaf_prefix = config.get("leaf_prefix", f"{site}-LEAF")
        bleaf_prefix = config.get("bleaf_prefix", f"{site}-BLEAF")
        bgw_prefix = config.get("bgw_prefix", f"{site}-BGW")
        mgmt_subnet = config.get("mgmt_subnet", "10.1.0")
        loopback_subnet = config.get("loopback_subnet", "10.1.255")
        vtep_subnet = config.get("vtep_subnet", "10.1.254")

        switches = hardware.get("switches", [])
        sfps = hardware.get("sfps", [])

        predominant_sfp = ""
        predominant_speed = "100G"
        if sfps:
            top_sfp = max(sfps, key=lambda s: s["quantity"])
            predominant_sfp = top_sfp["pid"]
            sfp_info = top_sfp.get("sfp_info", {})
            predominant_speed = sfp_info.get("speed", "100G")

        topology = BomParser._detect_topology_tier(switches)

        devices = []
        links = []
        sspine_idx = 0
        spine_idx = 0
        leaf_idx = 0
        bleaf_idx = 0
        bgw_idx = 0
        mgmt_ip_counter = 1

        for switch in switches:
            pid = switch["pid"]
            qty = switch["quantity"]
            role = switch.get("inferred_role", "leaf")
            role_hint = switch.get("role_hint", "").lower()

            if role_hint in ROLE_ALIASES:
                role = ROLE_ALIASES[role_hint]

            if topology["is_5_stage"] and role == "super_spine" and role_hint not in ("spine", "leaf"):
                pass
            elif topology["is_5_stage"] and role == "spine" and role_hint == "super_spine":
                role = "super_spine"
            elif topology["is_5_stage"] and role == "spine":
                model_info = NEXUS_MODELS.get(pid, {})
                if model_info.get("tier") == "super_spine" and role_hint not in ("spine",):
                    role = "super_spine"

            for i in range(qty):
                if role == "super_spine":
                    sspine_idx += 1
                    hostname = f"{sspine_prefix}-{sspine_idx:02d}"
                    lo0 = f"{loopback_subnet}.{200 + sspine_idx}/32"
                    lo1 = ""
                elif role == "spine":
                    spine_idx += 1
                    hostname = f"{spine_prefix}-{spine_idx:02d}"
                    lo0 = f"{loopback_subnet}.{spine_idx}/32"
                    lo1 = ""
                elif role == "border_leaf":
                    bleaf_idx += 1
                    hostname = f"{bleaf_prefix}-{bleaf_idx:02d}"
                    lo0 = f"{loopback_subnet}.{20 + bleaf_idx}/32"
                    lo1 = f"{vtep_subnet}.{20 + bleaf_idx}/32"
                elif role == "border_gateway":
                    bgw_idx += 1
                    hostname = f"{bgw_prefix}-{bgw_idx:02d}"
                    lo0 = f"{loopback_subnet}.{30 + bgw_idx}/32"
                    lo1 = f"{vtep_subnet}.{30 + bgw_idx}/32"
                else:
                    leaf_idx += 1
                    hostname = f"{leaf_prefix}-{leaf_idx:02d}"
                    lo0 = f"{loopback_subnet}.{10 + leaf_idx}/32"
                    lo1 = f"{vtep_subnet}.{10 + leaf_idx}/32"

                mgmt_ip_counter += 1
                device = {
                    "id": str(uuid.uuid4()),
                    "hostname": hostname,
                    "role": role,
                    "model": pid,
                    "serial": "",
                    "mgmt_ip": f"{mgmt_subnet}.{mgmt_ip_counter}/24",
                    "loopback0": lo0,
                    "loopback1": lo1,
                    "site": site,
                    "vpc_domain": "",
                    "vpc_peer": "",
                    "asn": "",
                    "interfaces": [],
                    "config": {},
                }
                devices.append(device)

        super_spine_devices = [d for d in devices if d["role"] == "super_spine"]
        spine_devices = [d for d in devices if d["role"] == "spine"]
        leaf_devices = [d for d in devices if d["role"] not in ("spine", "super_spine")]

        if topology["is_5_stage"] and super_spine_devices and spine_devices:
            spine_uplink_idx = 49
            for spine in spine_devices:
                for ss_num, sspine in enumerate(super_spine_devices, 1):
                    link = {
                        "id": str(uuid.uuid4()),
                        "from_device": sspine["hostname"],
                        "from_port": f"Ethernet1/{len([l for l in links if l['from_device'] == sspine['hostname']]) + 1}",
                        "to_device": spine["hostname"],
                        "to_port": f"Ethernet1/{spine_uplink_idx + ss_num - 1}",
                        "sfp": predominant_sfp,
                        "cable_type": "",
                        "speed": predominant_speed,
                    }
                    links.append(link)

            leaf_uplink_idx = 49
            for leaf in leaf_devices:
                for sp_num, spine in enumerate(spine_devices, 1):
                    link = {
                        "id": str(uuid.uuid4()),
                        "from_device": spine["hostname"],
                        "from_port": f"Ethernet1/{len([l for l in links if l['from_device'] == spine['hostname']]) + 1}",
                        "to_device": leaf["hostname"],
                        "to_port": f"Ethernet1/{leaf_uplink_idx + sp_num - 1}",
                        "sfp": predominant_sfp,
                        "cable_type": "",
                        "speed": predominant_speed,
                    }
                    links.append(link)
        else:
            uplink_port_idx = 49
            all_spine_like = super_spine_devices + spine_devices
            for leaf in leaf_devices:
                for sp_num, spine in enumerate(all_spine_like, 1):
                    link = {
                        "id": str(uuid.uuid4()),
                        "from_device": spine["hostname"],
                        "from_port": f"Ethernet1/{len([l for l in links if l['from_device'] == spine['hostname']]) + 1}",
                        "to_device": leaf["hostname"],
                        "to_port": f"Ethernet1/{uplink_port_idx + sp_num - 1}",
                        "sfp": predominant_sfp,
                        "cable_type": "",
                        "speed": predominant_speed,
                    }
                    links.append(link)

        for i in range(0, len(leaf_devices) - 1, 2):
            leaf_devices[i]["vpc_domain"] = str((i // 2) + 1)
            leaf_devices[i]["vpc_peer"] = leaf_devices[i + 1]["hostname"]
            leaf_devices[i + 1]["vpc_domain"] = str((i // 2) + 1)
            leaf_devices[i + 1]["vpc_peer"] = leaf_devices[i]["hostname"]

        sites = set(d.get("site", "") for d in devices if d.get("site"))
        return {
            "type": "fabric",
            "devices": devices,
            "links": links,
            "hardware": None,
            "metadata": {
                "total_devices": len(devices),
                "total_links": len(links),
                "sites": sorted(sites) if sites else ["site-1"],
                "multisite": len(sites) > 1,
                "bom_type": "generated_from_hardware",
                "topology": topology,
            }
        }

    @staticmethod
    def _detect_topology_tier(switches: list[dict]) -> dict:
        """
        Detect whether the fabric needs a 5-stage Clos (super-spine layer).
        Criteria:
        1. Multiple distinct spine-class models with different capabilities (400G vs 100G)
        2. Modular chassis present (N9K-C9508, C9516, C9408)
        3. Total leaf count > spine port capacity (need aggregation)
        4. High-radix spines (64p 400G) alongside standard spines (36p 100G)
        """
        spine_class = []
        leaf_class = []
        super_spine_candidates = []

        for sw in switches:
            pid = sw["pid"]
            role = sw.get("inferred_role", "leaf")
            model_info = NEXUS_MODELS.get(pid, {})
            tier = model_info.get("tier", "")
            qty = sw["quantity"]

            if role in ("spine", "super_spine") or tier in ("spine", "super_spine"):
                if tier == "super_spine" or model_info.get("modular") or model_info.get("ports_400g", 0) > 0:
                    super_spine_candidates.append({"pid": pid, "qty": qty, "model_info": model_info})
                else:
                    spine_class.append({"pid": pid, "qty": qty, "model_info": model_info})
            else:
                leaf_class.append({"pid": pid, "qty": qty, "model_info": model_info})

        total_spines = sum(s["qty"] for s in spine_class)
        total_super_spines = sum(s["qty"] for s in super_spine_candidates)
        total_leaves = sum(l["qty"] for l in leaf_class)

        has_distinct_spine_tiers = len(super_spine_candidates) > 0 and len(spine_class) > 0

        has_modular = any(s["model_info"].get("modular") for s in super_spine_candidates)

        max_spine_ports = max((s["model_info"].get("ports_100g", 36) for s in spine_class), default=36)
        spine_capacity_exceeded = total_leaves > (total_spines * max_spine_ports * 0.7) if total_spines > 0 else False

        is_5_stage = has_distinct_spine_tiers or has_modular or (spine_capacity_exceeded and total_super_spines > 0)

        reason = ""
        if is_5_stage:
            if has_modular:
                reason = "Modular chassis (N9K-C9508/9516/9408) detected as super-spine"
            elif has_distinct_spine_tiers:
                ss_pids = [s["pid"] for s in super_spine_candidates]
                sp_pids = [s["pid"] for s in spine_class]
                reason = f"Distinct spine tiers: super-spines ({', '.join(ss_pids)}) + spines ({', '.join(sp_pids)})"
            elif spine_capacity_exceeded:
                reason = f"Spine capacity exceeded: {total_leaves} leaves > {total_spines} spines * {max_spine_ports} ports"

        return {
            "is_5_stage": is_5_stage,
            "reason": reason,
            "super_spine_count": total_super_spines,
            "spine_count": total_spines,
            "leaf_count": total_leaves,
            "stages": 5 if is_5_stage else 3,
            "description": "5-Stage Clos (Super-Spine)" if is_5_stage else "3-Stage Clos (Spine-Leaf)",
        }

    @staticmethod
    def generate_template() -> bytes:
        """Generate a BOM Excel template for users."""
        wb = openpyxl.Workbook()

        ws_devices = wb.active
        ws_devices.title = "Devices"
        device_headers = [
            "Hostname", "Role", "Model", "Serial", "Mgmt IP",
            "Loopback0", "Loopback1 (VTEP)", "Site", "vPC Domain", "vPC Peer"
        ]
        ws_devices.append(device_headers)
        ws_devices.append([
            "DC1-SPINE-01", "spine", "N9K-C9336C-FX2", "FDO12345678",
            "10.1.0.1/24", "10.1.255.1/32", "", "DC1", "", ""
        ])
        ws_devices.append([
            "DC1-SPINE-02", "spine", "N9K-C9336C-FX2", "FDO12345679",
            "10.1.0.2/24", "10.1.255.2/32", "", "DC1", "", ""
        ])
        ws_devices.append([
            "DC1-LEAF-01", "leaf", "N9K-C93180YC-FX", "FDO22345678",
            "10.1.0.11/24", "10.1.255.11/32", "10.1.254.11/32", "DC1", "1", "DC1-LEAF-02"
        ])
        ws_devices.append([
            "DC1-LEAF-02", "leaf", "N9K-C93180YC-FX", "FDO22345679",
            "10.1.0.12/24", "10.1.255.12/32", "10.1.254.12/32", "DC1", "1", "DC1-LEAF-01"
        ])
        ws_devices.append([
            "DC1-BLEAF-01", "border_leaf", "N9K-C93180YC-FX", "FDO32345678",
            "10.1.0.21/24", "10.1.255.21/32", "10.1.254.21/32", "DC1", "2", "DC1-BLEAF-02"
        ])

        ws_cables = wb.create_sheet("Cabling")
        cable_headers = [
            "Hostname", "Local Port", "Connected To", "Remote Port",
            "SFP/Transceiver", "Cable Type", "Speed"
        ]
        ws_cables.append(cable_headers)
        ws_cables.append([
            "DC1-SPINE-01", "Ethernet1/1", "DC1-LEAF-01", "Ethernet1/49",
            "QSFP-100G-SR4", "MMF-OM4", "100G"
        ])
        ws_cables.append([
            "DC1-SPINE-01", "Ethernet1/2", "DC1-LEAF-02", "Ethernet1/49",
            "QSFP-100G-SR4", "MMF-OM4", "100G"
        ])

        ws_hw = wb.create_sheet("Hardware BOM")
        hw_headers = ["PID", "Description", "Qty", "Role (optional)"]
        ws_hw.append(hw_headers)
        ws_hw.append(["N9K-C9336C-FX2", "Nexus 9336C-FX2 36p 100G Spine", 2, "spine"])
        ws_hw.append(["N9K-C93180YC-FX", "Nexus 93180YC-FX 48p Leaf", 4, "leaf"])
        ws_hw.append(["N9K-C93180YC-FX", "Nexus 93180YC-FX 48p Border Leaf", 2, "border_leaf"])
        ws_hw.append(["QSFP-100G-SR4-S", "100G SR4 QSFP Transceiver", 24, ""])
        ws_hw.append(["FS-MM-OM4-10M", "OM4 MMF Patch Cable 10M", 24, ""])

        for ws in [ws_devices, ws_cables, ws_hw]:
            for col in ws.columns:
                max_length = max(len(str(cell.value or "")) for cell in col) + 2
                ws.column_dimensions[col[0].column_letter].width = max_length

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
